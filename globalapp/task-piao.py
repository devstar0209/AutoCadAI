import os
import re
import json
import cv2
import pytesseract
from pytesseract import Output
import openai
import openpyxl
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor
from pdf2image import convert_from_path
from PyPDF2 import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

# =================== CONFIG ===================
API_KEY = ""
client = openai.OpenAI(api_key=API_KEY)
pytesseract.pytesseract.tesseract_cmd = "/usr/bin/tesseract"  # Linux
# pytesseract.pytesseract.tesseract_cmd = r"C:\Path\To\tesseract.exe"  # Windows
MAX_WORKERS = 4

# CSI division-based keywords
construction_keywords = {
    "03 - Concrete": ["concrete", "slab", "footing", "foundation", "column", "beam", "girder",
                      "pile", "pier", "rebar", "reinforcement", "formwork", "joint", "curing"],
    "04 - Masonry": ["masonry", "brick", "block", "cmu", "stone", "veneer", "grout", "mortar", "lintel"],
    "05 - Metals": ["steel", "weld", "bolt", "plate", "angle", "channel", "pipe", "tube", "joist", "deck"],
    "06 - Wood": ["wood", "lumber", "timber", "plywood", "osb", "truss", "joist", "stud", "sheathing"],
    "07 - Thermal & Moisture": ["roof", "roofing", "membrane", "insulation", "vapor barrier",
                                "sealant", "flashing", "shingle", "tile"],
    "08 - Openings": ["door", "window", "frame", "glazing", "curtain wall", "skylight"],
    "09 - Finishes": ["floor", "ceiling", "tile", "carpet", "paint", "coating", "plaster",
                      "gypsum", "drywall", "veneer", "paneling"],
    "21 - Fire Protection": ["sprinkler", "fire protection", "standpipe", "fire pump", "alarm"],
    "22 - Plumbing": ["plumbing", "pipe", "valve", "toilet", "sink", "water heater", "drainage", "fixture"],
    "23 - HVAC": ["hvac", "duct", "chiller", "boiler", "air handler", "diffuser", "damper", "ventilation"],
    "26 - Electrical": ["electrical", "conduit", "cable", "wire", "panel", "transformer", "lighting",
                        "outlet", "switch", "breaker", "generator", "feeder", "grounding", "data", "telecom"],
    "Measurement Units": ["dimension", "length", "width", "height", "depth", "elevation", "level", "slope",
                          "thickness", "diameter", "radius", "area", "volume", "square", "cubic", "linear",
                          "feet", "foot", "inch", "inches", "meter", "millimeter", "centimeter"]
}

# =================== FRONTEND NOTIFY ===================
def notify_frontend(event_type, **kwargs):
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        "pdf_processing",
        {"type": event_type, **kwargs}
    )

# =================== OCR ===================
def extract_text_from_image(image_path: str) -> str:
    print(f"Entered OCR function: {image_path}")
    min_confidence = 40
    line_gap = 15
    try:
        if not os.path.exists(image_path):
            return ""

        img = cv2.imread(image_path)
        if img is None:
            print(f"[WARN] Failed to load image: {image_path}")
            return ""

        # === Preprocess image ===
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = cv2.bilateralFilter(gray, 9, 75, 75)  # edge-preserving denoise
        enhanced = cv2.convertScaleAbs(gray, alpha=1.5, beta=30)
        thresh = cv2.adaptiveThreshold(enhanced, 255,
                                    cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                    cv2.THRESH_BINARY, 11, 2)

        # === OCR with position data ===
        data = pytesseract.image_to_data(thresh, output_type=pytesseract.Output.DICT)

        lines = []
        current_line = []
        last_y = None

        for i, word in enumerate(data["text"]):
            word = word.strip()
            if not word:
                continue
            conf = int(data["conf"][i])
            if conf < min_confidence:
                continue

            y = data["top"][i]
            if last_y is None or abs(y - last_y) < line_gap:
                current_line.append(word)
            else:
                # new line detected
                lines.append(" ".join(current_line))
                current_line = [word]
            last_y = y

        if current_line:
            lines.append(" ".join(current_line))

        # === Post-cleaning ===
        cleaned_lines = []
        for line in lines:
            # print(f"line:: {line}")
            # remove single-character noise
            if len(line.strip()) < 2:
                continue
            # collapse extra spaces
            line = " ".join(line.split())
            cleaned_lines.append(line)

        structured_text = "\n".join(cleaned_lines)
        return structured_text.strip()

    except Exception as e:
        print(f"OCR error for {image_path}: {e}")
        return ""

# =================== AI COST ESTIMATION ===================


def should_use_nrm2(project_location=None, cad_text=""):
    """Determine if NRM2 standards should be used based on project characteristics"""

    # Caribbean countries that commonly use NRM2/RICS standards
    caribbean_countries = [
        "barbados", "trinidad", "tobago", "jamaica", "bahamas", "grenada",
        "st lucia", "dominica", "antigua", "barbuda", "st kitts", "nevis",
        "st vincent", "grenadines", "belize", "guyana", "suriname"
    ]

    # Commonwealth countries that may use NRM2
    commonwealth_indicators = [
        "commonwealth", "rics", "nrm2", "british", "uk standard",
        "metres", "cubic metres", "square metres"
    ]

    # Check project location
    if project_location:
        location_lower = project_location.lower()
        if any(country in location_lower for country in caribbean_countries):
            return True

    # Check CAD text for indicators
    # cad_lower = cad_text.lower()

    # # Look for metric units (strong indicator of NRM2)
    # metric_indicators = ["m³", "m²", "metres", "cubic metres", "square metres", "linear metres"]
    # if any(indicator in cad_lower for indicator in metric_indicators):
    #     return True

    # # Look for NRM2/RICS references
    # if any(indicator in cad_lower for indicator in commonwealth_indicators):
    #     return True

    # # Look for Caribbean location references in CAD text
    # if any(country in cad_lower for country in caribbean_countries):
    #     return True

    return False


def extract_project_location(cad_text):
    """Extract project location from CAD text to determine if NRM2 should be used"""
    if not cad_text:
        return None

    text_lower = cad_text.lower()

    # Caribbean countries and territories
    caribbean_locations = {
        "barbados": "Barbados",
        "trinidad": "Trinidad and Tobago",
        "tobago": "Trinidad and Tobago",
        "jamaica": "Jamaica",
        "bahamas": "Bahamas",
        "grenada": "Grenada",
        "st lucia": "Saint Lucia",
        "saint lucia": "Saint Lucia",
        "dominica": "Dominica",
        "antigua": "Antigua and Barbuda",
        "barbuda": "Antigua and Barbuda",
        "st kitts": "Saint Kitts and Nevis",
        "saint kitts": "Saint Kitts and Nevis",
        "nevis": "Saint Kitts and Nevis",
        "st vincent": "Saint Vincent and the Grenadines",
        "saint vincent": "Saint Vincent and the Grenadines",
        "grenadines": "Saint Vincent and the Grenadines",
        "belize": "Belize",
        "guyana": "Guyana",
        "suriname": "Suriname"
    }

    # Check for location indicators
    for location_key, location_name in caribbean_locations.items():
        if location_key in text_lower:
            return location_name

    # Check for other Commonwealth indicators
    commonwealth_indicators = ["commonwealth", "rics", "nrm2", "british standard"]
    for indicator in commonwealth_indicators:
        if indicator in text_lower:
            return "Commonwealth"

    return None


def get_construction_jobs(cad_text, project_location=None):
    print(f"Starting construction jobs analysis...")

    cad_text = preprocess_cad_text(cad_text)

    print(f"preprocess_cad_text: {cad_text}")

    # Determine if NRM2 standards should be used
    use_nrm2 = should_use_nrm2(project_location, cad_text)
    print(f"Using NRM2 standards: {use_nrm2}")

    # Build system prompt based on standards to use
    if use_nrm2:
        system_prompt = """You are a professional construction estimator with 20+ years of experience specializing in NRM2 (RICS) standards. Analyze CAD text and symbols to produce comprehensive, detailed cost estimates using NRM2 measurement principles.

1. NRM2 COMPLIANCE REQUIREMENTS:
   - Use NRM2 (New Rules of Measurement 2) - 2nd edition UK October 2021 - RICS standardized measurement rules
   - Apply net quantity measurement principles (measure to structural faces, exclude waste unless specified)
   - Follow RICS standard deduction rules for openings, overlaps, and double counting
   - Include 2024 MasterFormat (CSI) codes for international compatibility
   - Base costs on appropriate regional rates (Caribbean/Commonwealth markets)

2. DETAILED JOB ACTIVITY DESCRIPTIONS:
   Each Job Activity MUST include comprehensive specifications:
   - MATERIAL GRADE/TYPE: (e.g., "grade C25/30", "softwood stress graded C24", "PVC-U", "grade B500B")
   - DIMENSIONS/SIZES: (e.g., "300mm thick", "22mm diameter", "600x600mm", "2.5mm²")
        Apply default architectural and sitework/paving assumptions if dimensions are missing (see Tables below)
   - CONSTRUCTION METHOD: (e.g., "poured in-situ", "shop fabricated site erected", "bedded in cement mortar")
   - FINISH/QUALITY: (e.g., "smooth finish", "facework one side", "non-slip finish", "weather struck joint")
   - INCLUDED ITEMS: (e.g., "including vibrating and curing", "including ironmongery", "including terminations")

3. NRM2 WORK SECTIONS (MANDATORY):
   Use correct work section codes:
   - A10-A99: Preliminaries and General Conditions
   - D20-D99: Groundwork (excavation, filling, piling)
   - E10: In-situ concrete, E20: Formwork, E30: Reinforcement, E40: Precast concrete
   - F10: Brick/Block walling, F20: Natural stone, F30: Cast stone
   - G10: Structural steel, G20: Structural timber, G30: Metal decking
   - H10-H99: Cladding/Covering (curtain walling, sheeting, membranes)
   - H60: Roof coverings, H70: Rooflights
   - J10-J99: Waterproofing (liquid applied, sheet, cementitious)
   - K10-K99: Linings/Sheathing/Dry partitioning
   - L10: Windows/Rooflights, L20: Doors/Shutters, L30: Stairs/Walkways
   - M10-M99: Surface finishes (screeds, tiling, decorative papers, painting)
   - M20: Plastering/Rendering, M40: Stone/Concrete/Ceramic tiling, M50: Timber flooring
   - N10-N99: Furniture/Equipment
   - P10-P99: Building fabric sundries
   - Q10-Q99: Paving/Planting/Fencing/Site furniture
   - R10: Rainwater pipework/gutters, R11: Foul drainage above ground, R12: Drainage below ground
   - R13: Land drainage, R14: Sump and sewage pumping
   - S10-S99: Piped supply systems
   - S10: Cold water, S11: Hot water, S12: Central heating, S13: Compressed air, S14: Sanitary appliances/fittings
   - T10-T99: Mechanical heating/cooling/refrigeration systems
   - T30: Low temperature hot water heating, T31: Steam heating, T32: Warm air heating
   - T40: Cooling systems, T41: Chilled water, T42: Refrigeration
   - U10-U99: Ventilation/Air conditioning systems
   - U10: General ventilation, U11: Smoke extract/Smoke control, U12: Kitchen ventilation
   - U13: Car park ventilation, U14: Fume extract, U15: Air conditioning
   - V10-V99: Electrical supply/power/lighting systems
   - V10: Electrical generation plant, V11: HV supply/distribution, V12: LV supply/distribution
   - V13: Motors/Starters, V20: LV distribution, V21: General LV power, V22: General lighting
   - V23: Emergency lighting, V24: Specialist lighting, V25: Exit/Emergency signs
   - W10-W99: Communications/Security/Control systems
   - W10: Telecommunications, W20: Radio/TV/CCTV/Video, W21: Data transmission
   - W22: Audio systems, W23: Clocks, W30: Security systems, W40: Building management systems
   - X10-X99: Transport systems (lifts, escalators, moving walkways)
   - Y10-Y99: General fixings/supports/restraints
   - Z10-Z99: Simple building works incidental to landscape works

4. REQUIRED OUTPUT FORMAT:
   - EXACT JSON list only: [{"CSI code":"03 30 00","NRM2 Section":"E10","Category":"In-situ Concrete","Job Activity":"Reinforced concrete foundation, grade C25/30, poured in-situ, 300mm thick, including vibrating and curing","Quantity":25,"Unit":"m³","Rate":185.50,"Material Cost":2787.50,"Equipment Cost":464.00,"Labor Cost":1391.00,"Total Cost":4642.50}]
   - All fields mandatory; numeric values for costs/quantities/rate
   - "NRM2 Section" field with work section code (A10-Z99)
   - Units MUST be metric: m³, m², m, nr, kg, tonnes
   - Steel reinforcement Quantity and UNIT in Kilograms
   - Structural Steel to be Measured in "Tonnes"

5. DETAILED SPECIFICATION EXAMPLES:
   CONCRETE (E10): "Reinforced concrete foundation, grade C25/30, poured in-situ, 300mm thick, including vibrating and curing"
   MASONRY (F10): "Common brickwork, 215mm thick, in cement mortar 1:3, English bond, facework one side, pointed with weather struck joint"
   STEEL (G10): "Structural steel beams, grade S355, universal beam 356x171x67kg/m, shop fabricated, site erected with bolted connections"
   ELECTRICAL DISTRIBUTION (V20): "PVC insulated copper cable, 2.5mm², single core, drawn into 25mm PVC conduit, including terminations"
   ELECTRICAL POWER (V21): "13A socket outlets, twin switched with earth terminal, including back box and connections"
   ELECTRICAL LIGHTING (V22): "LED ceiling luminaires, 600x600mm, 36W, 4000K, including lamp, control gear and mounting accessories"
   PLUMBING (S10): "Copper pipes, 22mm diameter, table X, capillary fittings, including brackets and pipe insulation where required"
   HVAC (U15): "Air conditioning split unit, 5kW cooling capacity, wall mounted, including refrigerant pipework and controls"
   DRAINAGE (R12): "Vitrified clay drain pipes, 150mm diameter, flexible joints, laid in trenches including granular bedding"

6. COVERAGE AND NAME PRESERVATION (CRITICAL):
   - Use CAD_OCR_TEXT as ground truth. Do not omit major categories with strong signals.
   - PRESERVE EXACT ITEM NAMES from CAD_OCR_TEXT for named equipment/devices (NO sequential generation or pattern-based extrapolation).
   - CRITICAL: If CAD_OCR_TEXT contains "Panel EDL216-2", use ONLY "Panel EDL216-2" - do NOT create EDL216-3, EDL216-4, etc.
   - PROPAGATE ATTRIBUTES into the Job Activity text when present in CAD_OCR_TEXT:
       • rating (e.g., 125A, 30 kVA, 250 kW)
       • size (e.g., 22mm, 600x600mm, 2.5mm²)
       • length (from patterns like L=200, L=4", L=3.5m; treat as quantity + unit where possible)
   - PRESERVE MULTIPLICITY per distinct name/size/rating: separate rows and correct quantities for each distinct item.

   - For each drawing, you MUST extract and include ALL relevant construction elements if present, including but not limited to:
       • Foundations (strip, pad, raft, pile, etc.)
       • Walls (internal, external, retaining, block, brick, concrete, etc.)
       • Slabs (ground, upper, suspended, etc.)
       • Columns, Beams, Girders
       • Stairs, Walkstairs, Landings, Ramps
       • Roofs, Roof coverings, Rooflights
       • Doors, Windows, Openings
       • Finishes (floor, wall, ceiling, tiling, painting, etc.)
       • Sitework (paving, kerbs, landscaping, fencing, etc.)
       • Drainage, Plumbing, HVAC, Electrical, Fire Protection, Equipment
   - Do NOT skip or lose any items. If an item is present in the CAD_OCR_TEXT, it must be included in the output.
   - ANTI-HALLUCINATION: Use ONLY items that exist in CAD_OCR_TEXT - do not generate additional similar items.


6. COVERAGE AND PRESERVATION:
    - Use CAD_OCR_TEXT as ground truth
    - PRESERVE exact item names, ratings, sizes from CAD text (NO modifications or sequential generation)
    - Create separate line items for distinct specifications
    - Include ONLY items that actually exist in the CAD_OCR_TEXT
    - DO NOT generate similar items based on patterns (e.g., if EDL216-2 exists, do NOT create EDL216-3, EDL216-4)

7. DERIVATION RULES FOR MISSING ITEMS:
   - Branch cable (m) = (outlets + switches + fixtures) × 4m average
   - Feeder cable (m) = (panels + major equipment) × 60m average
   - Conduit (m) ≈ 9% of total cable length
   - Formwork (m²) = concrete contact area
   - Reinforcement = 80-120 kg per m³ of concrete
   - Apply Default Architectural and Sitework/Paving Assumptions only if CAD/OCR text is missing dimensions/specifications.

   **Default Architectural Assumptions Table**
   | Element | Typical / Default Size / Specification | Notes |
   |---------|----------------------------------------|-------|
   | Room Sizes | Bedroom (single): 9–12 m² | Floor area if missing |
   | | Bedroom (double): 12–18 m² | Floor area if missing |
   | | Living room: 12–30 m² | Small to large units |
   | | Dining room: 8–20 m² | Floor area if missing |
   | | Bathroom: 3–6 m² | Floor area if missing |
   | Ceiling Height | 2.4–2.7 m | Standard between finished floor and ceiling |
   | Wall Thickness | Internal walls: 100–150 mm | Brick/blockwork |
   | | External walls: 200–300 mm | Brick/blockwork or concrete |
   | Doors | Internal: 2.1 m × 0.8 m | Standard flush/panel |
   | | External: 2.1 m × 0.9 m | Includes frame |
   | Windows | Standard: 1.2 m × 1.0 m | Single/double glazing |
   | | Large / patio: 2.1 m × 1.8 m | Sliding/fixed |
   | Floor / Slab Thickness | Ground floor slab: 150–200 mm RC | Reinforced concrete incl. screed |
   | | Upper floor slab: 120–180 mm RC | Reinforced concrete incl. screed |
   | Stairs | Rise: 150–180 mm | Vertical height per step |
   | | Tread: 250–300 mm | Horizontal depth per step |
   | | Width: 900–1200 mm | Clear width of stair |
   | | Landings: 900–1200 mm | Minimum clear width/depth |
   | Doors & Openings Deduction | Deduct from wall/floor areas as per NRM2 | Only if specified in CAD/OCR |
   | Default Construction Method | Concrete: poured in-situ; masonry: bedded in cement mortar | Apply unless CAD text specifies otherwise |
   | Default Finish | Walls: smooth plaster/render; Floors: screed/tiled; Ceilings: painted | Apply unless CAD text specifies otherwise |

   **Default Sitework / Paving Assumptions Table**
   | Element | Typical / Default Size / Specification | Notes |
   |---------|----------------------------------------|-------|
   | Concrete Paving | 100–150 mm thick reinforced concrete slab | Standard pedestrian/light vehicle traffic |
   | Asphalt / Bituminous Pavement | 50–100 mm thick compacted asphalt | For driveways / parking areas |
   | Kerbs | 150 mm high × 200 mm wide | Standard concrete kerbs |
   | Footpaths / Walkways | 1.0–1.5 m width | Pedestrian access, concrete or pavers |
   | Driveways | 3–4 m width | Single lane, reinforced or compacted |
   | Site Furniture Spacing | Benches: 2–3 m apart; Trash bins: 15–20 m apart | Default if CAD does not specify |
   | Landscaping / Planting Beds | 1–3 m width | Shrubs or groundcover, default if CAD missing |
   | Fencing | 1.8–2 m high | Chain link, timber, or metal standard |
   | Paving Unit Sizes | Standard pavers: 200 × 100 × 60 mm | Modular concrete or clay pavers |
   | Expansion / Joint Spacing | Concrete paving: 4–6 m centers | Joints for shrinkage and thermal movement |

8. NRM2 COST STRUCTURE:
   - Total Cost = Material Cost + Labor Cost + Equipment Cost
   - Rate = Material Rate
   - If only the material rate is known, first calculate Total Cost and Material Cost as (Material Rate × Quantity), then derive Labor and Equipment costs using standard allocation percentages.
   - Material 60-70%, Labor 25-35%, Equipment 5-15%
   - Include regional adjustments for Caribbean markets

9. QUALITY VALIDATION:
   - Net quantities measured correctly
   - Appropriate work sections assigned
   - Detailed specifications included
   - Metric units consistently used
   - Cost relationships balanced
"""
    else:
        system_prompt = """
        You are a professional construction estimator with 20+ years of experience. Analyze CAD text and symbols to produce comprehensive cost estimates.

1. ACCURACY REQUIREMENTS:
   - Use 2024 MasterFormat (CSI) codes.
   - Electrical labor: Use 2023–2024 NECA standards.
   - Base costs on current US market rates (RSMeans-like, national averages).
   - Consider regional variation (use national average if unknown).
   - Quantities must be realistic for the described scope and scale.

2. REQUIRED OUTPUT FORMAT:
   - Output EXACT JSON list only, no prose:
     [{"CSI code":"03 30 00","Category":"Concrete","Job Activity":"Cast-in-place Concrete Slab, 6-inch thick","Quantity":150,"Unit":"CY","Rate":125.5,"Material Cost":12000,"Equipment Cost":3000,"Labor Cost":4500,"Total Cost":19500}]
   - All fields are mandatory; costs/quantities/rates must be numeric.
   - CSI codes must be in progressive six-digit sequence (with decimal extensions if needed).
   - Job Activity must be specific and self-contained. Each activity is a separate line item.

3. CATEGORY KEYWORDS:
   - CONCRETE: slab, paving, footing, column, beam, wall, sidewalk, driveway, footing, capping beam, raft, tie beam, transfer slab
   - MASONRY: brick, CMU, block, stone, veneer, grout, mortar, lintel
   - METALS: structural steel, beam, column, stair, handrail, joist, decking, weld, bolt, plate, angle, channel, pipe, tube
   - WOOD: lumber, timber, plywood, OSB, truss, joist, stud, sheathing
   - THERMAL & MOISTURE: roofing, membrane, insulation, vapor barrier, sealant, flashing, shingle, tile
   - OPENINGS: door, window, frame, glazing, curtain wall, skylight
   - FINISHES: flooring, ceiling, tile, carpet, paint, coating, plaster, gypsum, drywall, veneer, paneling
   - FIRE PROTECTION: sprinkler, standpipe, fire pump, alarm, fire system
   - PLUMBING: pipe, valve, toilet, sink, water heater, fixture, drainage
   - HVAC: duct, chiller, boiler, air handler, diffuser, damper, ventilation, fan, AC
   - ELECTRICAL: conduit, cable, wire, panel, switchboard, transformer (or transf), generator, lighting, fixture, outlet, switch, breaker, receptacle, dimmer
   - COMMUNICATIONS: data, telecom
   - ELECTRONIC SAFETY: CCTV, security, alarm
   - EARTHWORK / SITEWORK: excavation, grading, backfill, asphalt, curb, landscaping, paving, sidewalk, driveway
   - EQUIPMENT: elevator, lift

4. EXTRACTION RULES:
   - Scan **every line in CAD_OCR_TEXT** and assign it to the proper category using the keywords above.
   - Every detected item must produce a JSON line item.
   - Preserve exact names, sizes, ratings, and multiplicities from CAD_OCR_TEXT; do not generate additional items.
   - If size, thickness, or rating is missing, use **DEFAULT IMPERIAL ASSUMPTIONS**.

5. QUANTITY/RATE/COST RULES:
   - Default Cost Allocation if breakdown unknown:
       • Material: 50–65%
       • Labor: 30–40%
       • Equipment: 5–15%
   - Total Cost Formula:
       • Material Cost = Rate × Quantity
       • Total Cost = Material + Labor + Equipment
   - Use RSMeans 2024 national average for installed costs; electrical labor follows NECA 2023–2024.
   - If OCR provides only material rate, backfill labor/equipment per above allocation.
   - If OCR provides total installed cost, split according to above bands.

6. CONCRETE & REBAR:
    - This ensures that “Paving concrete” is recognized as category 03 CONCRETE → paving.
   - All concrete items (slabs, footings, beams, columns, walls, paving, sidewalks, driveways, etc.) **must include a separate rebar line item**.
   - Concrete Paving:
       • Unit = SF
       • Rebar = 1.46 - 2.81 lb/SF
   - Other concrete items:
       • Unit = CY
       • Rebar = typical reinforcement density per element type (see Section 9)
   - Rebar JSON line item format: "Reinforcing Steel #4 rebar in [Concrete Item Name]"
   - Rebar must always be a separate JSON object, not embedded in concrete description.

7. ELECTRICAL ITEMS:
   - Include Branch cable (2#12 AWG + 1#12 GRD) if Lighting, receptacles, switches and dimmers(or dim) are present. quantity is calculated as (total number of devices) × 12 LF (or 25 LF if long runs implied).
   - Include Feeder cable (2#10 AWG + 1#10 GRD) if panels, switchboards (Amp), transformers (kVA) and generators (kW) are present. quantity is calculated as (total numbers of items) × 200 LF 
   - Output **one line item for 3/4"C conduit if Branch cable (2#12 AWG + 1#12 GRD) or Feeder cable (2#10 AWG + 1#10 GRD) are present**; quantity = 0.3 × (Branch cable (2#12 AWG + 1#12 GRD) LF + Feeder cable (2#10 AWG + 1#10 GRD) LF).

8. DEFAULT IMPERIAL ASSUMPTIONS:
   - Use these defaults when CAD/OCR text does not specify sizes/thicknesses:
     **Architectural Defaults**
       * Room Sizes: Bedroom (single) 97–130 ft², Bedroom (double) 130–194 ft², Living 130–323 ft², Dining 86–215 ft², Bathroom 32–65 ft²
       * Ceiling Height: 7.9–8.9 ft
       * Wall Thickness: Internal 4–6 in, External 8–12 in
       * Doors: Internal 6'10"×2'7", External 6'10"×2'11"
       * Windows: Standard 4'×3'3", Large/Patio 6'10"×5'11"
       * Floor/Slab Thickness: Ground 6–8 in RC, Upper 4.7–7 in RC
       * Stairs: Rise 6–7 in, Tread 10–12 in, Width 3–4 ft, Landings 3–4 ft
       * Default Construction: Concrete poured in-situ; masonry bedded in cement mortar
       * Default Finish: Walls smooth plaster/render; Floors screed/tiled; Ceilings painted

     **Sitework / Paving Defaults**
       * Foundation / Footings: 24–36 in depth, 12–36 in width (typical residential/commercial)
       * Concrete Paving: 4–6 in thick reinforced concrete slab
       * Asphalt: 2–4 in thick compacted
       * Kerbs: 6 in high × 8 in wide
       * Footpaths: 3–5 ft width
       * Driveways: 10–13 ft width
       * Landscaping Beds: 3–10 ft width
       * Fencing: 6–6.5 ft high
       * Paving Units: 8×4×2.4 in standard
       * Expansion / Joint Spacing: Concrete paving 13–20 ft centers

9. CONCRETE REINFORCEMENT RATES (LB/FT³):
   | Element             | Weight (lb/ft³) | Steel Share (%) |
   |---------------------|-----------------|----------------|
   | Bases               | 5.6–8.1         | 1.2–1.7 %      |
   | Beams               | 15.6–21.8       | 3.2–4.5 %      |
   | Capping Beams       | 8.4             | 1.7 %          |
   | Columns             | 12.4–28         | 2.5–5.7 %      |
   | Ground Beams        | 14.3–20.6       | 2.9–4.2 %      |
   | Footings            | 4.3–6.2         | 0.9–4.2 %      |
   | Pile Caps           | 6.8–9.3         | 1.4–1.9 %      |
   | Plate Slabs         | 5.9–8.4         | 1.2–1.7 %      |
   | Rafts               | 7.1             | 1.5 %          |
   | Retaining Walls     | 6.8–9.3         | 1.4–1.9 %      |
   | Ribbed Floor Slabs  | 5–7.4           | 1–1.5 %        |
   | Slabs – One Way     | 4.6–7.8         | 1–1.6 %        |
   | Slabs – Two Way     | 4.1–8.4         | 0.9–1.7 %      |
   | Stairs              | 8.1–10.6        | 1.7–2.2 %      |
   | Tie Beams           | 8.1–10.6        | 1.7–2.2 %      |
   | Transfer Slabs      | 9.3             | 1.9 %          |
   | Walls – Normal      | 4.3–6.2         | 0.9–1.3 %      |
   | Walls – Wind        | 5.6–9.3         | 1.1–1.9 %      |
   - Use these to calculate rebar quantities for all concrete elements.
   - Rebar output must always be a separate JSON object with proper CSI code and description.

10. VALIDATION CHECKLIST:
   - Cover all categories present in CAD_OCR_TEXT.
   - Positive quantities/costs; math consistent.
   - Preserve exact item names, sizes, ratings, multiplicities.
   - Anti-hallucination: Only extract what exists in CAD_OCR_TEXT.
   - Output JSON only, one object per item, including concrete rebar and electrical cables.

"""

    user_prompt = f"""CAD_OCR_TEXT: 
{cad_text}

Instruction: Extract every construction item from the CAD text into a valid JSON array following all rules in the system prompt. Include:

- All detected items per category using Section 3 keywords.
- Concrete items including paving, slabs, footings, walls, beams, columns, etc.
- Separate rebar line items for all concrete elements.
- Electrical items including dimmers, lighting, receptacles, panels, transformers, switchboards, generators.
- Branch/Feeder cables and conduit if applicable.
- Apply DEFAULT IMPERIAL ASSUMPTIONS when size, thickness, or rating is missing.
- Preserve exact item names, multiplicity, and attributes (kVA, kW, Amp, sizes, etc.)
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[    
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1,
        )
        response_text = response.choices[0].message.content
        print(f"Initial AI response received:: {response_text}")
        validated_response = validate_and_improve_response(response_text, cad_text)
        return validated_response
    except Exception as e:
        print(f"Error in get_construction_jobs: {e}")
        return get_construction_jobs_fallback(cad_text)


def validate_and_improve_response(response_text, cad_text):
    """Validate AI response and improve if necessary"""
    try:
        # Try to extract and parse JSON
        clean_json = extract_json_from_response(response_text)
        if not clean_json:
            print("No JSON found in response, attempting to extract...")
            # Try to find JSON without code blocks
            import re
            json_match = re.search(r'\[.*\]', response_text, re.DOTALL)
            if json_match:
                clean_json = json_match.group(0)

        if clean_json:
            parsed_data = json.loads(clean_json)

            # Validate the data structure and content
            validation_result = validate_construction_data(parsed_data)

            if validation_result["is_valid"]:
                print(f"Response validation passed: {validation_result['score']}/100")
                return parsed_data
            else:
                print(f"Response validation failed: {validation_result['errors']}")
                # Try to improve the response
                return improve_response_with_feedback(response_text, validation_result, cad_text)
        else:
            print("Could not extract valid JSON, using fallback...")
            return get_construction_jobs_fallback(cad_text)

    except Exception as e:
        print(f"Validation error: {e}")
        return get_construction_jobs_fallback(cad_text)

def validate_construction_data(data):
    """Validate construction data for accuracy and completeness"""
    errors = []
    score = 100

    if not isinstance(data, list):
        errors.append("Data is not a list")
        return {"is_valid": False, "errors": errors, "score": 0}

    if len(data) == 0:
        errors.append("No construction activities found")
        return {"is_valid": False, "errors": errors, "score": 0}

    required_fields = ["CSI code", "Category", "Job Activity", "Quantity", "Unit", "Rate", "Material Cost", "Equipment Cost", "Labor Cost", "Total Cost"]
    # NRM2 Section is optional but validated if present
    optional_nrm2_fields = ["NRM2 Section"]

    for i, item in enumerate(data):
        if not isinstance(item, dict):
            errors.append(f"Item {i} is not a dictionary")
            score -= 20
            continue

        # Check required fields
        for field in required_fields:
            if field not in item:
                errors.append(f"Item {i} missing field: {field}")
                score -= 5

        # Validate numeric fields
        numeric_fields = ["Quantity", "Rate", "Material Cost", "Equipment Cost", "Labor Cost", "Total Cost"]
        for field in numeric_fields:
            if field in item:
                try:
                    value = float(item[field])
                    if value <= 0:
                        errors.append(f"Item {i} {field} must be positive")
                        score -= 3
                except (ValueError, TypeError):
                    errors.append(f"Item {i} {field} is not a valid number")
                    score -= 5

        # Validate cost relationships
        if all(field in item for field in ["Material Cost", "Equipment Cost", "Labor Cost", "Total Cost"]):
            try:
                calculated_total = float(item["Material Cost"]) + float(item["Equipment Cost"]) + float(item["Labor Cost"])
                actual_total = float(item["Total Cost"])
                if abs(calculated_total - actual_total) > 0.01:  # Allow small rounding differences
                    errors.append(f"Item {i} total cost doesn't match sum of components")
                    score -= 10
            except (ValueError, TypeError):
                pass

        # Validate CSI code format (allow flexibility for NRM2 integration)
        if "CSI code" in item:
            csi_code = str(item["CSI code"])
            # Basic CSI format validation - allow some flexibility
            if not re.match(r'^\d{2}[\s\-]?\d{2}[\s\-]?\d{2}', csi_code):
                errors.append(f"Item {i} has invalid CSI code format: {csi_code}")
                score -= 3

        # Validate NRM2 section format if present
        if "NRM2 Section" in item:
            nrm2_section = str(item["NRM2 Section"])
            if not re.match(r'^[A-Z]\d{2}', nrm2_section):
                errors.append(f"Item {i} has invalid NRM2 section format: {nrm2_section}")
                score -= 3

        # Validate metric units for NRM2 compliance
        if "Unit" in item:
            unit = str(item["Unit"]).lower()
            metric_units = ["m³", "m²", "m", "nr", "kg", "tonnes", "m3", "m2"]
            imperial_units = ["cy", "sf", "lf", "ea", "tons", "cf"]

            # Check if using NRM2 (has NRM2 Section field or metric units)
            has_nrm2_section = "NRM2 Section" in item
            uses_metric = any(mu in unit for mu in metric_units)

            if has_nrm2_section and not uses_metric:
                errors.append(f"Item {i} has NRM2 section but non-metric unit: {unit}")
                score -= 2

    is_valid = score >= 70 and len(errors) <= 3
    return {"is_valid": is_valid, "errors": errors, "score": score}

def improve_response_with_feedback(original_response, validation_result, cad_text):
    """Improve AI response based on validation feedback"""
    feedback_prompt = f"""The previous response had these issues: {validation_result['errors']}

Please provide a corrected response that addresses these problems. Focus on:
1. Ensuring all required fields are present
2. Making sure all costs are positive numbers
3. Verifying that Total Cost = Material Cost + Equipment Cost + Labor Cost
4. Using proper CSI code format (XX XX XX)
5. Providing realistic quantities and costs

Original CAD text: {cad_text}

Return ONLY a valid JSON array with the exact structure specified in the system prompt."""

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are a professional construction estimator. Fix the previous response based on the feedback provided."},
                {"role": "user", "content": feedback_prompt}
            ],
            temperature=0.1,
            max_tokens=4000
        )

        improved_response = response.choices[0].message.content
        validated_response = validate_and_improve_response(improved_response, cad_text)

        print("Improved response generated based on feedback")
        return validated_response

    except Exception as e:
        print(f"Error improving response: {e}")
        return get_construction_jobs_fallback(cad_text)

def get_construction_jobs_fallback(cad_text):
    """Fallback method with simpler, more reliable approach"""
    print("Using fallback method for construction jobs...")

    fallback_prompt = f"""Based on this CAD drawing text, provide a simple list of common construction activities with basic cost estimates:

{cad_text}

Return a JSON array with this exact format:
[{{"CSI code": "03 30 00", "Category": "Concrete", "Job Activity": "Foundation", "Quantity": 100, "Unit": "CY", "Rate": 120, "Material Cost": 8000, "Equipment Cost": 2000, "Labor Cost": 2000, "Total Cost": 12000}}]

Include only the most obvious construction activities. Keep it simple and accurate."""

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a construction estimator. Provide simple, accurate cost estimates."},
                    {"role": "user", "content": fallback_prompt}
                ],
                temperature=0.1,
                max_tokens=2000
            )

        return response.choices[0].message.content

    except Exception as e:
        print(f"Fallback method also failed: {e}")
        # Return error message instead of default data
        return f'{{"error": "AI analysis failed: {str(e)}", "message": "Unable to process CAD drawing for cost estimation. Please check your OpenAI API key and try again."}}'

# =================== PDF PAGE PROCESSING ===================
def convert_pdf_page_to_image(pdf_path: str, page_number: int) -> str:
    images = convert_from_path(pdf_path, dpi=200, first_page=page_number, last_page=page_number, poppler_path="/usr/bin")
    if not images: return ""
    directory = os.path.dirname(pdf_path)
    image_path = os.path.join(directory, f"page_{page_number}.png")
    images[0].save(image_path, "PNG")
    return image_path

# =================== EXCEL OUTPUT GENERATION ===================
def generate_outputs(output_json: dict, filename: str):
    print(f"Generating Excel output...")
    """
    Generate Excel file with cost estimation in the required format:
    - Column A: Di (Division number)
    - Column B: Description
    - Column C: Main Building (cost)
    - Column D: Total (cost)
    
    Returns: Excel file (.xlsx) only, no PDF
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Summary"
    
    # Define styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    summary_font = Font(bold=True)
    summary_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header row
    ws.append(["", "CSI code", "Description", "Main Building", "Total"])
    
    # Style header
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Track totals
    facility_subtotal = 0
    facility_items = []
    
    # Ordered categories to mirror the Summary sheet exactly
    ordered_categories = [
        ("Existing Conditions", "02"),
        ("Concrete", "03"),
        ("Masonry", "04"),
        ("Wood, Plastic & Composites", "06"),
        ("Thermal & Moisture Protection", "07"),
        ("Openings", "08"),
        ("Finishes", "09"),
        ("Specialties", "10"),
        ("Plumbing", "22"),
        ("HVAC", "23"),
        ("Electrical", "26"),
        ("Electronic Safety & Security", "28"),
        ("Earthwork", "31"),
    ]
    csi_division_map = {name: div for name, div in ordered_categories}

    def normalize_category(raw: str) -> str:
        text = (raw or "").lower().strip()
        # First check exact matches
        exact_matches = {
            "existing conditions": "Existing Conditions",
            "concrete": "Concrete",
            "masonry": "Masonry",
            "wood, plastic & composites": "Wood, Plastic & Composites",
            "thermal & moisture protection": "Thermal & Moisture Protection",
            "openings": "Openings",
            "finishes": "Finishes",
            "specialties": "Specialties",
            "plumbing": "Plumbing",
            "hvac": "HVAC",
            "electrical": "Electrical",
            "electronic safety & security": "Electronic Safety & Security",
            "earthwork": "Earthwork",
        }
        if text in exact_matches:
            return exact_matches[text]
        
        # Then check partial matches (order matters - more specific first)
        # Check for concrete-related keywords that might be misclassified
        concrete_keywords = ["concrete", "slab", "footing", "foundation", "column", "beam", "wall", "sidewalk", "driveway", "paving", "capping beam", "raft", "tie beam", "transfer slab"]
        if any(keyword in text for keyword in concrete_keywords):
            # Only if it's not clearly earthwork
            if not any(ew_word in text for ew_word in ["excavation", "grading", "backfill", "earth moving"]):
                return "Concrete"
        
        mapping = [
            ("concrete", "Concrete"),  # Check concrete first before other matches
            ("existing conditions", "Existing Conditions"),
            ("existing", "Existing Conditions"),
            ("masonry", "Masonry"),
            ("brick", "Masonry"),
            ("block", "Masonry"),
            ("cmu", "Masonry"),
            ("wood", "Wood, Plastic & Composites"),
            ("lumber", "Wood, Plastic & Composites"),
            ("timber", "Wood, Plastic & Composites"),
            ("plastic", "Wood, Plastic & Composites"),
            ("composite", "Wood, Plastic & Composites"),
            ("thermal", "Thermal & Moisture Protection"),
            ("moisture", "Thermal & Moisture Protection"),
            ("roof", "Thermal & Moisture Protection"),
            ("roofing", "Thermal & Moisture Protection"),
            ("opening", "Openings"),
            ("door", "Openings"),
            ("window", "Openings"),
            ("finish", "Finishes"),
            ("finishes", "Finishes"),
            ("special", "Specialties"),
            ("specialties", "Specialties"),
            ("plumb", "Plumbing"),
            ("plumbing", "Plumbing"),
            ("hvac", "HVAC"),
            ("mechanical", "HVAC"),
            ("electric", "Electrical"),
            ("electrical", "Electrical"),
            ("electronic safety", "Electronic Safety & Security"),
            ("security", "Electronic Safety & Security"),
            ("earth", "Earthwork"),
            ("earthwork", "Earthwork"),
            ("sitework", "Earthwork"),
            ("excavation", "Earthwork"),
            ("grading", "Earthwork"),
        ]
        for key, value in mapping:
            if key in text:
                return value
        return raw or "Uncategorized"
    
    def get_div_for_category(category_name: str) -> str:
        return csi_division_map.get(category_name, "02")
    
    def generate_csi_code(category: str, job_activity: str = "") -> str:
        """Generate appropriate CSI code based on category and job activity"""
        cat_lower = category.lower()
        job_lower = (job_activity or "").lower()
        
        # Existing Conditions
        if "existing" in cat_lower:
            return "02 00 00"
        
        # Concrete - check this early to catch concrete items
        if "concrete" in cat_lower:
            if any(word in job_lower for word in ["slab", "foundation", "footing", "column", "beam", "wall"]):
                return "03 30 00"  # Cast-in-place Concrete
            elif "rebar" in job_lower or "reinforcement" in job_lower or "reinforcing" in job_lower:
                return "03 20 00"  # Concrete Reinforcing
            elif "formwork" in job_lower or "form" in job_lower:
                return "03 11 00"  # Concrete Forming
            elif "finish" in job_lower:
                return "03 35 00"  # Concrete Finishing
            else:
                return "03 30 00"  # Default: Cast-in-place Concrete
        
        # Masonry
        if "masonry" in cat_lower or "brick" in cat_lower or "block" in cat_lower or "cmu" in cat_lower:
            if "brick" in job_lower:
                return "04 21 00"  # Clay Unit Masonry
            elif "block" in job_lower or "cmu" in job_lower:
                return "04 22 00"  # Concrete Unit Masonry
            else:
                return "04 20 00"  # Unit Masonry
        
        # Metals / Steel
        if "metal" in cat_lower or "steel" in cat_lower:
            if "deck" in job_lower:
                return "05 15 00"  # Metal Decking
            else:
                return "05 12 00"  # Structural Steel Framing
        
        # Wood
        if "wood" in cat_lower or "lumber" in cat_lower or "timber" in cat_lower:
            if "sheathing" in job_lower:
                return "06 16 00"  # Sheathing
            else:
                return "06 10 00"  # Rough Carpentry
        
        # Thermal & Moisture Protection
        if "thermal" in cat_lower or "moisture" in cat_lower or "roof" in cat_lower:
            if "insulation" in job_lower:
                return "07 21 00"  # Building Insulation
            elif "vapor" in job_lower or "barrier" in job_lower:
                return "07 26 00"  # Vapor Retarders
            else:
                return "07 20 00"  # Thermal Insulation
        
        # Openings
        if "opening" in cat_lower or "door" in cat_lower or "window" in cat_lower:
            if "door" in job_lower:
                if "metal" in job_lower:
                    return "08 11 00"  # Metal Doors and Frames
                else:
                    return "08 14 00"  # Wood Doors
            elif "window" in job_lower:
                return "08 51 00"  # Windows
            else:
                return "08 10 00"  # Openings
        
        # Finishes
        if "finish" in cat_lower:
            if "gypsum" in job_lower or "drywall" in job_lower:
                return "09 21 00"  # Gypsum Board
            elif "plaster" in job_lower:
                return "09 29 00"  # Gypsum Plaster
            elif "tile" in job_lower:
                return "09 30 00"  # Tiling
            elif "floor" in job_lower or "carpet" in job_lower:
                return "09 65 00"  # Resilient Flooring
            else:
                return "09 20 00"  # Plaster and Gypsum Board
        
        # Specialties
        if "special" in cat_lower:
            return "10 00 00"  # Specialties
        
        # Fire Protection
        if "fire" in cat_lower or "sprinkler" in job_lower:
            return "21 13 00"  # Fire-Suppression Sprinkler Systems
        
        # Plumbing
        if "plumb" in cat_lower:
            if "fixture" in job_lower or "toilet" in job_lower or "sink" in job_lower:
                return "22 11 00"  # Plumbing Fixtures
            elif "water" in job_lower or "distribution" in job_lower:
                return "22 13 00"  # Facility Water Distribution
            elif "sewer" in job_lower or "drainage" in job_lower:
                return "22 16 00"  # Facility Sanitary Sewerage
            else:
                return "22 10 00"  # Plumbing
        
        # HVAC
        if "hvac" in cat_lower or "mechanical" in cat_lower:
            if "duct" in job_lower:
                return "23 21 00"  # HVAC Ducts and Casings
            elif "coil" in job_lower:
                return "23 23 00"  # HVAC Coils
            elif "air" in job_lower or "distribution" in job_lower:
                return "23 25 00"  # HVAC Air Distribution
            else:
                return "23 20 00"  # HVAC
        
        # Electrical
        if "electric" in cat_lower:
            if "conduit" in job_lower or "raceway" in job_lower:
                return "26 28 00"  # Low-Voltage Electrical Power Raceways
            elif "cable" in job_lower or "conductor" in job_lower or "wire" in job_lower:
                return "26 27 00"  # Low-Voltage Electrical Power Conductors
            elif "panel" in job_lower or "switchboard" in job_lower or "transformer" in job_lower:
                return "26 20 00"  # Low-Voltage Electrical Power Transmission
            elif "light" in job_lower or "fixture" in job_lower:
                return "26 51 00"  # Interior Lighting
            else:
                return "26 05 00"  # Common Work Results for Electrical
        
        # Electronic Safety & Security
        if "security" in cat_lower or "electronic" in cat_lower:
            if "access" in job_lower or "control" in job_lower:
                return "28 13 00"  # Access Control
            elif "intrusion" in job_lower or "detection" in job_lower or "alarm" in job_lower:
                return "28 16 00"  # Intrusion Detection
            else:
                return "28 10 00"  # Electronic Access Control and Intrusion Detection
        
        # Earthwork
        if "earth" in cat_lower or "sitework" in cat_lower:
            if "excavation" in job_lower or "fill" in job_lower:
                return "31 23 00"  # Excavation and Fill
            elif "earth" in job_lower or "moving" in job_lower:
                return "31 25 00"  # Earth Moving
            elif "paving" in job_lower or "base" in job_lower or "ballast" in job_lower:
                return "31 32 00"  # Bases, Ballasts, and Paving
            else:
                return "31 20 00"  # Earth Moving
        
        # Default based on division
        division = get_div_for_category(category)
        return f"{division} 00 00"
    
    # Add Facility Costs
    ws.append(["", "", "Facility:", "", ""])
    facility_row = ws.max_row
    for col in range(1, 6):
        cell = ws.cell(row=facility_row, column=col)
        cell.border = border_style
    
    # Process items from output_json
    details = output_json.get("Details", [])
    summary = output_json.get("Summary", [])
    
    # Only filter out items that are truly invalid (no description/category AND no cost)
    # Keep items even if all costs are 0, as long as they have a job/category (visibility required)
    valid_details = []
    for item in details:
        total_cost = float(item.get("Total Cost", 0) or 0)
        material = float(item.get("Material Cost", 0) or 0)
        labor = float(item.get("Labor Cost", 0) or 0)
        equipment = float(item.get("Equipment Cost", 0) or 0)
        has_identity = bool(str(item.get("Job Activity", "")).strip() or str(item.get("Category", "")).strip())
        # Exclude only if no identity fields AND all costs are zero
        if not has_identity and total_cost == 0 and material == 0 and labor == 0 and equipment == 0:
            continue
        valid_details.append(item)
    details = valid_details
    
    # Aggregate totals by Category for the summary sheet (one row per category)
    category_totals = {}
    category_csi_codes = {}  # Map category to CSI code
    categories_present = set()
    
    # Debug: Track original category names to see what we're getting
    original_categories = {}
    
    for item in details:
        original_cat = str(item.get("Category", "Uncategorized"))
        job_activity = str(item.get("Job Activity", "")).lower()
        
        # Re-classify based on job activity if category seems wrong
        # Check for concrete-related keywords in job activity
        # IMPORTANT: Only reclassify if clearly misclassified - don't reclassify "Existing Conditions" items
        cat = normalize_category(original_cat)
        
        # Only reclassify if category is clearly wrong (e.g., electrical or uncategorized with concrete keywords)
        # Do NOT reclassify "Existing Conditions" - those are legitimate even if they mention concrete
        concrete_keywords = ["patch", "slab", "footing", "foundation", "column", "beam", "concrete", 
                           "dowel", "turnstile slab", "slab-on-grade", "cast-in-place", "poured"]
        if any(keyword in job_activity for keyword in concrete_keywords):
            # Only reclassify if clearly misclassified (not Existing Conditions, not already Concrete)
            if "existing" not in original_cat.lower() and "concrete" not in original_cat.lower():
                # Check if it's a clear misclassification (electrical, uncategorized, etc.)
                if "electrical" in original_cat.lower() or original_cat.lower() in ["uncategorized", "other"]:
                    cat = "Concrete"
        
        # Track original category names for debugging
        if original_cat not in original_categories:
            original_categories[original_cat] = cat
        
        try:
            subtotal = float(item.get("Total Cost", 0) or 0)
        except Exception:
            subtotal = 0.0
        
        # Process all items, even if cost is 0 (they might have partial data)
        # Only exclude if truly empty (all zeros)
        material_check = float(item.get("Material Cost", 0) or 0)
        labor_check = float(item.get("Labor Cost", 0) or 0)
        equipment_check = float(item.get("Equipment Cost", 0) or 0)
        
        # Track presence regardless of costs to ensure category visibility
        categories_present.add(cat)

        if subtotal > 0 or material_check > 0 or labor_check > 0 or equipment_check > 0:
            category_totals[cat] = category_totals.get(cat, 0.0) + subtotal

        # Extract or generate CSI code from item even if subtotal is zero (for display rows)
        if cat not in category_csi_codes:
            # For Concrete category, always generate correct CSI code (03 30 00) instead of trusting item's code
            # This ensures Concrete gets the right code even if items were misclassified
            if cat == "Concrete":
                job_activity_name = item.get("Job Activity", "")
                generated_csi = generate_csi_code(cat, job_activity_name)
                category_csi_codes[cat] = generated_csi
            else:
                csi_code = item.get("CSI code", "")
                if csi_code:
                    # Normalize CSI code format: remove spaces and ensure 6-8 digits
                    csi_clean = re.sub(r'[\s\-]', '', str(csi_code))
                    if len(csi_clean) >= 6:
                        # Verify the CSI code is appropriate for the category
                        # Check if first two digits match expected division
                        expected_division = get_div_for_category(cat)
                        if csi_clean[:2] == expected_division:
                            # Format as XX XX XX or XX XX XX XX (with spaces)
                            if len(csi_clean) == 6:
                                formatted_csi = f"{csi_clean[0:2]} {csi_clean[2:4]} {csi_clean[4:6]}"
                            else:
                                formatted_csi = f"{csi_clean[0:2]} {csi_clean[2:4]} {csi_clean[4:6]} {csi_clean[6:]}"
                            category_csi_codes[cat] = formatted_csi
                        else:
                            # CSI code doesn't match category, generate correct one
                            job_activity_name = item.get("Job Activity", "")
                            generated_csi = generate_csi_code(cat, job_activity_name)
                            category_csi_codes[cat] = generated_csi
                    else:
                        # Invalid CSI code format, generate correct one
                        job_activity_name = item.get("Job Activity", "")
                        generated_csi = generate_csi_code(cat, job_activity_name)
                        category_csi_codes[cat] = generated_csi
                else:
                    # Generate CSI code if not present
                    job_activity_name = item.get("Job Activity", "")
                    generated_csi = generate_csi_code(cat, job_activity_name)
                    category_csi_codes[cat] = generated_csi
    
    # Debug output
    print(f"📊 Category totals: {category_totals}")
    print(f"📊 Original categories mapped: {original_categories}")
    
    def get_csi_code_for_category(category_name: str) -> str:
        """Get CSI code for category, defaulting to division-based code if not found"""
        if category_name in category_csi_codes:
            return category_csi_codes[category_name]
        # Default CSI code based on division (format: XX 00 00)
        division = get_div_for_category(category_name)
        return f"{division} 00 00"

    # Write categories in the exact order from ordered_categories
    # Include categories that have items (even if total is 0, for completeness)
    row_num = ws.max_row + 1
    for cat, division in ordered_categories:
        total_for_cat = category_totals.get(cat, 0.0)
        # Write category if it is present in details, even if total is 0
        if cat in categories_present:
            csi_code = get_csi_code_for_category(cat)
            ws.append([division, csi_code, cat, total_for_cat, total_for_cat])
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border_style
                if col in (4, 5):
                    cell.number_format = '#,##0.00'
            facility_subtotal += total_for_cat
            row_num = ws.max_row + 1
    
    # Add Subtotal Facility
    ws.append(["", "", "Subtotal Facility", facility_subtotal, facility_subtotal])
    subtotal_row = ws.max_row
    
    for col in range(1, 6):
        cell = ws.cell(row=subtotal_row, column=col)
        cell.font = summary_font
        cell.border = Border(
            top=Side(style='double'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        if col == 4 or col == 5:
            cell.number_format = '#,##0.00'
    
    # Calculate GC Markups (typically 6% of facility subtotal)
    gc_markups = facility_subtotal * 0.06
    ws.append(["", "", "GC Markups", gc_markups, gc_markups])
    gc_row = ws.max_row
    
    for col in range(1, 6):
        cell = ws.cell(row=gc_row, column=col)
        cell.border = border_style
        if col == 4 or col == 5:
            cell.number_format = '#,##0.00'
    
    # Calculate Total Construction Cost
    total_construction = facility_subtotal + gc_markups
    ws.append(["", "", "Total Construction Cost", total_construction, total_construction])
    const_row = ws.max_row
    
    for col in range(1, 6):
        cell = ws.cell(row=const_row, column=col)
        cell.font = summary_font
        cell.border = Border(
            top=Side(style='double'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        if col == 4 or col == 5:
            cell.number_format = '#,##0.00'
    
    # Calculate Contingency (typically 5% of total construction)
    contingency = total_construction * 0.05
    ws.append(["", "", "Contingency", contingency, contingency])
    cont_row = ws.max_row
    
    for col in range(1, 6):
        cell = ws.cell(row=cont_row, column=col)
        cell.border = border_style
        if col == 4 or col == 5:
            cell.number_format = '#,##0.00'
    
    # Calculate Total Project Cost
    total_project = total_construction + contingency
    ws.append(["", "", "Total Project Cost", total_project, total_project])
    total_row = ws.max_row
    
    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = Border(
            top=Side(style='double'),
            bottom=Side(style='double'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        if col == 4 or col == 5:
            cell.number_format = '$#,##0'
    
    # Add Area metrics (calculate from facility details if available)
    # This would need to be calculated from the CAD drawings
    area = 1608  # Default or calculate from items
    cost_per_sf = total_project / area if area > 0 else 0
    
    ws.append(["", "", "Area", area, ""])
    ws.append(["", "", "$/sf", round(cost_per_sf, 0), ""])
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    # =================== MAIN BUILDING TAB ===================
    main_ws = wb.create_sheet(title="Main Building")
    # Detailed columns per category (to reflect your example)
    main_headers = [
        "Div", "CSI code", "DESCRIPTION", "Quant.", "Unit", "U/Cost", "Cost", "Total.Material",
        "MH/Unit", "Hrs.", "Rate", "Total.Labor",
        "H/Unit", "Hrs.", "Rate", "Total.Equip",
        "Markups", "Total Amount"
    ]
    main_ws.append(main_headers)

    for col in range(1, len(main_headers) + 1):
        cell = main_ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = Alignment(horizontal='center', vertical='center')

    details = output_json.get("Details", [])
    
    # Only filter out items that are truly invalid (no description/category AND no cost)
    # Keep items even if all costs are 0, as long as they have a job/category (visibility required)
    valid_details = []
    for item in details:
        total_cost = float(item.get("Total Cost", 0) or 0)
        material = float(item.get("Material Cost", 0) or 0)
        labor = float(item.get("Labor Cost", 0) or 0)
        equipment = float(item.get("Equipment Cost", 0) or 0)
        has_identity = bool(str(item.get("Job Activity", "")).strip() or str(item.get("Category", "")).strip())
        if not has_identity and total_cost == 0 and material == 0 and labor == 0 and equipment == 0:
            continue
        valid_details.append(item)
    details = valid_details
    
    # Create category_csi_codes map for Main Building tab (reuse from summary processing)
    category_csi_codes_main = {}
    for item in details:
        original_cat = str(item.get("Category", "Uncategorized"))
        job_activity = str(item.get("Job Activity", "")).lower()
        
        # Re-classify based on job activity if category seems wrong
        # IMPORTANT: Only reclassify if clearly misclassified - preserve Existing Conditions, Specialties, etc.
        cat = normalize_category(original_cat)
        
        # Only reclassify if category is clearly wrong (e.g., electrical or uncategorized with concrete keywords)
        # Do NOT reclassify "Existing Conditions", "Specialties", "Electronic Safety & Security" - those are legitimate
        concrete_keywords = ["patch", "slab", "footing", "foundation", "column", "beam", "concrete", 
                           "dowel", "turnstile slab", "slab-on-grade", "cast-in-place", "poured", "paving", "sidewalk", "driveway"]
        if any(keyword in job_activity for keyword in concrete_keywords):
            # Only reclassify if clearly misclassified (not Existing Conditions, not already Concrete, not Specialties, etc.)
            if "existing" not in original_cat.lower() and "concrete" not in original_cat.lower() and "special" not in original_cat.lower():
                # Check if it's a clear misclassification (electrical, uncategorized, etc.)
                if "electrical" in original_cat.lower() or original_cat.lower() in ["uncategorized", "other"]:
                    cat = "Concrete"
        
        if cat not in category_csi_codes_main:
            # For Concrete category, always generate correct CSI code (03 30 00) instead of trusting item's code
            if cat == "Concrete":
                category_csi_codes_main[cat] = generate_csi_code(cat, job_activity)
            else:
                csi_code = item.get("CSI code", "")
                if csi_code:
                    csi_clean = re.sub(r'[\s\-]', '', str(csi_code))
                    if len(csi_clean) >= 6:
                        # Verify the CSI code is appropriate for the category
                        expected_division = get_div_for_category(cat)
                        if csi_clean[:2] == expected_division:
                            if len(csi_clean) == 6:
                                category_csi_codes_main[cat] = f"{csi_clean[0:2]} {csi_clean[2:4]} {csi_clean[4:6]}"
                            else:
                                category_csi_codes_main[cat] = f"{csi_clean[0:2]} {csi_clean[2:4]} {csi_clean[4:6]} {csi_clean[6:]}"
                        else:
                            # CSI code doesn't match category, generate correct one
                            category_csi_codes_main[cat] = generate_csi_code(cat, job_activity)
                    else:
                        category_csi_codes_main[cat] = generate_csi_code(cat, job_activity)
                else:
                    category_csi_codes_main[cat] = generate_csi_code(cat, job_activity)

    # Default hourly rates for labor and equipment per category
    labor_hourly_rate = {
        "Existing Conditions": 95.0,
        "Concrete": 95.0,
        "Masonry": 95.0,
        "Wood, Plastic & Composites": 115.0,
        "Thermal & Moisture Protection": 120.0,
        "Openings": 120.0,
        "Finishes": 150.0,
        "Specialties": 150.0,
        "Plumbing": 220.0,
        "HVAC": 180.0,
        "Electrical": 165.0,
        "Electronic Safety & Security": 165.0,
        "Earthwork": 120.0,
    }
    equipment_hourly_rate = 125.0
    
    # Group details by Category to create sections like the example
    # Re-classify items based on job activity if category is wrong
    category_to_items = {}
    for it in details:
        original_cat = str(it.get("Category", "Uncategorized"))
        job_activity = str(it.get("Job Activity", "")).lower()
        
        # Re-classify based on job activity if category seems wrong
        # IMPORTANT: Only reclassify if clearly misclassified - preserve Existing Conditions, Specialties, etc.
        cat = normalize_category(original_cat)
        
        # Only reclassify if category is clearly wrong (e.g., electrical or uncategorized with concrete keywords)
        # Do NOT reclassify "Existing Conditions", "Specialties", "Electronic Safety & Security" - those are legitimate
        concrete_keywords = ["patch", "slab", "footing", "foundation", "column", "beam", "concrete", 
                           "dowel", "turnstile slab", "slab-on-grade", "cast-in-place", "poured", "paving", "sidewalk", "driveway"]
        if any(keyword in job_activity for keyword in concrete_keywords):
            # Only reclassify if clearly misclassified (not Existing Conditions, not already Concrete, not Specialties, etc.)
            if "existing" not in original_cat.lower() and "concrete" not in original_cat.lower() and "special" not in original_cat.lower():
                # Check if it's a clear misclassification (electrical, uncategorized, etc.)
                if "electrical" in original_cat.lower() or original_cat.lower() in ["uncategorized", "other"]:
                    cat = "Concrete"
        
        category_to_items.setdefault(cat, []).append(it)

    # Write high-level section title 'FACILITY:'
    main_ws.append(["", "", "FACILITY:", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    fac_row = main_ws.max_row
    for col in range(1, len(main_headers) + 1):
        c = main_ws.cell(row=fac_row, column=col)
        c.font = summary_font
        c.border = border_style
    
    def get_csi_code_for_category_main_building(category_name: str) -> str:
        """Get CSI code for category in Main Building tab"""
        # Try to get from category_csi_codes_main if available, otherwise generate
        if category_name in category_csi_codes_main:
            return category_csi_codes_main[category_name]
        # Generate based on category
        division = get_div_for_category(category_name)
        return f"{division} 00 00"
    
    # Order categories by division where possible to keep a sensible order
    def sort_key(cat: str):
        return get_div_for_category(cat) + cat

    main_total = 0.0
    # Use the same order as the Summary tab
    for category, _div in ordered_categories:
        if category not in category_to_items:
            continue
        division_for_category = get_div_for_category(category)
        # Get CSI code for this category
        csi_code_for_category = get_csi_code_for_category_main_building(category)
        
        # Category header, bold row
        main_ws.append([division_for_category, csi_code_for_category, category, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]) 
        cat_header_row = main_ws.max_row
        for col in range(1, len(main_headers) + 1):
            ch = main_ws.cell(row=cat_header_row, column=col)
            ch.font = summary_font
            ch.border = border_style

        category_subtotal = 0.0
        for item in category_to_items[category]:
            subtotal = float(item.get("Total Cost", 0) or 0)
            # Include items even if total is 0, as long as they have some data
            material = float(item.get("Material Cost", 0) or 0)
            labor = float(item.get("Labor Cost", 0) or 0)
            equipment = float(item.get("Equipment Cost", 0) or 0)
            # Only skip if truly empty
            if subtotal == 0 and material == 0 and labor == 0 and equipment == 0:
                continue
                
            description = item.get("Job Activity", "")
            try:
                qty = float(item.get("Quantity") or 0)
            except Exception:
                qty = 0.0
            unit = item.get("Unit", "")
            material = float(item.get("Material Cost") or 0)
            labor = float(item.get("Labor Cost") or 0)
            equipment = float(item.get("Equipment Cost") or 0)
            sub_markups = float(item.get("Sub Markups") or 0)

            # Calculate material unit cost: Material Cost / Quantity
            # This reflects current market prices for the specific material
            if qty > 0:
                material_unit_cost = material / qty
            else:
                # Fallback to Rate if quantity is 0 or missing
                material_unit_cost = float(item.get("Rate") or 0)
            unit_cost = material_unit_cost
            
            # Get CSI code for this item
            item_csi_code = item.get("CSI code", "")
            if item_csi_code:
                # Normalize CSI code format
                csi_clean = re.sub(r'[\s\-]', '', str(item_csi_code))
                if len(csi_clean) >= 6:
                    if len(csi_clean) == 6:
                        item_csi_code = f"{csi_clean[0:2]} {csi_clean[2:4]} {csi_clean[4:6]}"
                    else:
                        item_csi_code = f"{csi_clean[0:2]} {csi_clean[2:4]} {csi_clean[4:6]} {csi_clean[6:]}"
                else:
                    item_csi_code = csi_code_for_category
            else:
                # Generate CSI code if not present
                item_csi_code = generate_csi_code(category, description)
            
            # Labor block
            lh_rate = labor_hourly_rate.get(category, 150.0)
            mh_unit = ((labor / lh_rate) / (qty if qty else 1.0)) if lh_rate else 0.0
            labor_hours = mh_unit * qty
            # Equipment block
            eh_rate = equipment_hourly_rate
            eh_unit = ((equipment / eh_rate) / (qty if qty else 1.0)) if eh_rate else 0.0
            equip_hours = eh_unit * qty

            main_ws.append([
                division_for_category, item_csi_code, description, qty, unit,
                unit_cost, subtotal, material,
                mh_unit, labor_hours, lh_rate, labor,
                eh_unit, equip_hours, eh_rate, equipment,
                sub_markups, sub_markups
            ])
            # Style numeric cells for this detail row
            row_idx = main_ws.max_row
            for col in range(1, len(main_headers) + 1):
                cell = main_ws.cell(row=row_idx, column=col)
                cell.border = border_style
                if col >= 5:
                    cell.number_format = '#,##0.00'

            category_subtotal += subtotal
            main_total += subtotal

        # Category subtotal row
        main_ws.append(["", "", f"Subtotal: {category}", "", "", "", "", "", "", "", "", "", "", "", "", "", "", category_subtotal])
        sr = main_ws.max_row
        for col in range(1, len(main_headers) + 1):
            cell = main_ws.cell(row=sr, column=col)
            cell.font = summary_font
            cell.border = border_style
            if col == len(main_headers):
                cell.number_format = '#,##0.00'

    # Total row
    main_ws.append(["", "", "Total", "", "", "", "", "", "", "", "", "", "", "", "", "", "", main_total])
    total_row = main_ws.max_row
    for col in range(1, len(main_headers) + 1):
        cell = main_ws.cell(row=total_row, column=col)
        cell.font = summary_font
        cell.border = border_style
        if col >= 5:
            cell.number_format = '#,##0.00'

    # Column widths
    main_ws.column_dimensions['A'].width = 6
    main_ws.column_dimensions['B'].width = 12
    main_ws.column_dimensions['C'].width = 60
    for col in ['D','E']:
        main_ws.column_dimensions[col].width = 10
    for col in ['F','G','H','I','J','K','L','M','N','O','P','Q']:
        main_ws.column_dimensions[col].width = 14

    # (Labor Rates tab removed)

    # Save Excel
    wb.save(filename)
    print(f"✅ Exported estimate to {filename}")

# =================== SUMMARY GENERATION ===================
def generate_summary_from_details(details: list) -> dict:
    """
    Generate a cost summary grouped by Category from the Details list.
    Returns a dict: { "Summary": [...], "Details": details }
    """
    if not details:
        return {"Summary": [], "Details": []}

    summary_map = {}
    total_project_cost = 0.0

    for item in details:
        category = item.get("Category", "Uncategorized").strip()
        try:
            # Try "Total Cost" first, fallback to "Subtotal Cost" for compatibility
            subtotal = float(item.get("Total Cost", 0))
            if subtotal == 0:
                subtotal = float(item.get("Subtotal Cost", 0))
        except (TypeError, ValueError):
            subtotal = 0.0

        summary_map[category] = summary_map.get(category, 0.0) + subtotal
        total_project_cost += subtotal

    summary_list = []
    for category, total in summary_map.items():
        summary_list.append({
            "Category": category,
            "Total Cost": round(total, 2)
        })

    summary_list.append({
        "Category": "Total Project Cost",
        "Total Cost": round(total_project_cost, 2)
    })

    return {
        "Summary": summary_list,
        "Details": details
    }

# =================== MAIN PDF PROCESSING WITH LIVE PROGRESS ===================
def get_page_count(pdf_file):
    reader = PdfReader(pdf_file)
    return len(reader.pages)
def start_pdf_processing(pdf_path: str, output_pdf: str, output_excel: str, location=None):
    total_pages = get_page_count(pdf_path)
    # all_results = []  # store combined structured outputs

    # def process_page(page_num):
    #     img_path = convert_pdf_page_to_image(pdf_path, page_num)
    #     if not img_path:
    #         return ""

    #     # OCR for this page
    #     page_text = extract_text_from_image(img_path)
    #     if not page_text.strip():
    #         return ""

    #     print(f"🔹 Running AI extraction on page {page_num}/{total_pages}...")
    #     try:
    #         # Pass per-page text to AI (your existing fine-tuned function)
    #         page_result = get_construction_jobs(page_text, location)

    #         # Parse and validate JSON
    #         if isinstance(page_result, str):
    #             try:
    #                 page_result = json.loads(page_result)
    #             except json.JSONDecodeError:
    #                 print(f"⚠️ Invalid JSON on page {page_num}, skipping.")
    #                 return ""

    #         # Append valid structured results
    #         if isinstance(page_result, list):
    #             all_results.extend(page_result)

    #     except Exception as e:
    #         print(f"⚠️ Error processing page {page_num}: {e}")

    #     # Notify frontend
    #     progress = round((page_num / total_pages) * 100, 2)
    #     notify_frontend(
    #         "page_processed",
    #         page=page_num,
    #         total_pages=total_pages,
    #         progress=progress
    #     )
    all_texts = [""] * total_pages

    def process_page(page_num):
        img_path = convert_pdf_page_to_image(pdf_path, page_num)
        if img_path:
            all_texts[page_num-1] = extract_text_from_image(img_path)
        progress = round((page_num / total_pages) * 100, 2)
        notify_frontend(
            "page_processed",
            page=page_num,
            total_pages=total_pages,
            progress=progress
        )

    # 🧵 Run pages concurrently
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        executor.map(process_page, range(1, total_pages + 1))

    combined_text = " ".join(all_texts)
    # Try to extract project location from PDF metadata or text
    # project_location = extract_project_location(combined_text)
    jobs_list = get_construction_jobs(combined_text, location)
    
    # --- Safety guards so generate_outputs never sees malformed data ---
    if isinstance(jobs_list, str):
        try:
            parsed = json.loads(jobs_list)
            jobs_list = parsed if isinstance(parsed, list) else []
        except json.JSONDecodeError:
            print("⚠️ jobs_list returned invalid JSON string, defaulting to empty list.")
            jobs_list = []
    elif not isinstance(jobs_list, list):
        print(f"⚠️ jobs_list returned unsupported type {type(jobs_list)}, defaulting to empty list.")
        jobs_list = []

    safe_jobs = []
    for idx, item in enumerate(jobs_list):
        if isinstance(item, dict):
            safe_jobs.append(item)
            continue
        if isinstance(item, str):
            try:
                parsed_item = json.loads(item)
                if isinstance(parsed_item, dict):
                    safe_jobs.append(parsed_item)
                elif isinstance(parsed_item, list):
                    safe_jobs.extend(obj for obj in parsed_item if isinstance(obj, dict))
            except json.JSONDecodeError:
                print(f"⚠️ Skipping malformed string entry at index {idx}")
        else:
            print(f"⚠️ Skipping unsupported entry at index {idx}: {type(item)}")

    if safe_jobs:
        final_output = generate_summary_from_details(safe_jobs)
        generate_outputs(final_output, output_excel)
        notify_frontend(
            "pdf_processing_completed",
            pdf_path=output_pdf,
            excel_path=output_excel,
            progress=100
        )

def preprocess_cad_text(cad_text: str) -> str:
    """
    Preprocess CAD text to preserve multi-line item relationships while cleaning for AI processing.
    - Preserves line breaks for multi-line items
    - Groups related lines together
    - Cleans excessive whitespace while maintaining structure
    """

    # Split into lines and clean each line
    lines = cad_text.split('\n')
    cleaned_lines = []

    for line in lines:
        # Remove excessive spaces but preserve single spaces
        line = re.sub(r'\s+', ' ', line.strip())
        # Skip empty lines
        if line:
            cleaned_lines.append(line.upper())

    # Join with newlines to preserve multi-line structure
    processed_text = '\n'.join(cleaned_lines)

    # Group related multi-line items more intelligently
    lines = processed_text.split('\n')
    grouped_items = []
    current_item_lines = []

    for i, line in enumerate(lines):
        # Check if this line looks like a continuation or a new item
        is_short_line = len(line.split()) <= 3  # Short lines might be continuations
        has_partial_word = not line.endswith(' ') and i < len(lines) - 1  # Line ends abruptly
        next_line_exists = i < len(lines) - 1

        # If current group exists and this line seems like a continuation
        if current_item_lines and (is_short_line or has_partial_word):
            current_item_lines.append(line)
        else:
            # Save current group if it exists
            if current_item_lines:
                grouped_items.append(' '.join(current_item_lines))
                current_item_lines = []

            # Start new group
            current_item_lines = [line]

    # Add the last group
    if current_item_lines:
        grouped_items.append(' '.join(current_item_lines))

    # Clean up the grouped items
    final_items = []
    for item in grouped_items:
        # Remove extra spaces and clean up
        item = re.sub(r'\s+', ' ', item.strip())
        if item:
            final_items.append(item)

    return '\n\n'.join(final_items)

def extract_json_from_response(response_text: str):
    match = re.search(r"```json\n(.*?)\n```", response_text, re.DOTALL)
    if match: return match.group(1)
    match2 = re.search(r'\[.*\]', response_text, re.DOTALL)
    return match2.group(0) if match2 else None

def validate_data(data: list) -> bool:
    return isinstance(data, list) and len(data) > 0