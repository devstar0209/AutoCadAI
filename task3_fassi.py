import os
import re
import json
import math
import time
import difflib
import pickle
from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple
from difflib import SequenceMatcher
from collections import defaultdict
import numpy as np
from functools import lru_cache
import faiss

# OCR / PDF
from pdf2image import convert_from_path
import pytesseract

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

from openai import OpenAI

from dotenv import load_dotenv

load_dotenv()



# -----------------------------
# USER CONFIG
# -----------------------------
OPENAI_API_KEY=os.getenv("OPENAI_API_KEY")
client = OpenAI(api_key=OPENAI_API_KEY)
MODEL = "gpt-5.2"

PRICE_JSON = "resources_enriched.json"
CACHE_DIR = "cache-fassi"
CANDIDATES_JSON = os.path.join(CACHE_DIR, "csi_candidates.json")
FAISS_PKL = os.path.join(CACHE_DIR, "csi_faiss.pkl")

FAISS_THRESHOLD = 0.6
TOP_K = 3

# -------------------------------
# MODELS
# -------------------------------
MATCH_MODEL = SentenceTransformer("sentence-transformers/all-MiniLM-L6-v2", cache_folder="./models")


ALLOWED_CATEGORIES = [
    "General Requirements",
    "Existing Conditions",
    "Concrete",
    "Masonry",
    "Metal",
    "Wood, Plastics, and Composites",
    "Thermal and Moisture Protection",
    "Openings",
    "Finishes",
    "Specialties",
    "Equipment",
    "Furnishing",
    "Special Construction",
    "Conveying Systems",
    "Fire Suppression",
    "Plumbing",
    "HVAC",
    "Electrical",
    "Communications",
    "Electronic Safety & Security",
    "Earthwork",
    "Exterior Improvement",
    "Utilities",
    "Transportations",
    "Waterway & marine",
    "Material Processing & Handling Equipment",
    "Pollution Control Equipment"
]

# =====================================================
# SYSTEM ONTOLOGY (CRITICAL)
# =====================================================

SYSTEM_ONTOLOGY = {

    "General Conditions": [
        "general conditions", "temporary facilities", "mobilization",
        "supervision", "project management", "permits"
    ],

    "Excavation & Earthwork System": [
        "earthwork", "excavation", "grading", "trenching",
        "backfill", "compaction", "cut and fill", "dewatering"
    ],

    "Concrete Structure System": [
        "cast-in-place", "foundation", "footing",
        "slab", "structural concrete", "rebar", "reinforcing"
    ],

    "Masonry System": [
        "masonry", "cmu", "concrete block", "brick",
        "stone", "mortar", "grout"
    ],

    "Structural Steel System": [
        "structural steel", "steel framing", "steel beam",
        "steel column", "metal deck", "joist"
    ],

    "Wood Framing System": [
        "wood framing", "stud", "joist", "truss",
        "sheathing", "plywood", "osb", "blocking"
    ],

    "Roofing System": [
        "roofing", "roof membrane", "shingle",
        "tpo", "epdm", "bur", "flashing", "roof insulation"
    ],

    "Exterior Envelope System": [
        "exterior wall", "facade", "cladding",
        "curtain wall", "eifs", "siding", "air barrier"
    ],

    "Interior Finishes System": [
        "interior finishes", "drywall", "gypsum board",
        "partition", "ceiling", "flooring", "paint"
    ],

    "Doors and Windows System": [
        "door", "window", "frame", "hardware",
        "storefront", "glazing", "louvers", "skylight"
    ],

    "Plumbing System": [
        "plumbing", "domestic water", "water supply",
        "hot water", "cold water", "plumbing fixtures", "valve"
    ],

    "Sanitary Sewer System": [
        "sanitary sewer", "sanitary drainage", "waste piping",
        "soil pipe", "vent piping", "cleanout"
    ],

    "Manhole and Structure": [
        "manhole", "mh#", "mh ", "access structure",
        "catch basin", "inspection chamber",
        "precast manhole", "utility vault"
    ],

    "Storm Drainage System": [
        "storm drainage", "storm sewer", "roof drain",
        "area drain", "catch basin", "downspout", "leader"
    ],

    "HVAC System": [
        "hvac", "mechanical", "heating", "cooling",
        "ventilation", "air handling", "ductwork", "exhaust"
    ],

    "Electrical System": [
        "electrical", "power", "lighting",
        "panel", "feeder", "branch circuit", "grounding"
    ],

    "Fire Protection System": [
        "fire protection", "sprinkler", "standpipe",
        "fire pump", "fire riser"
    ],

    "Low Voltage / Communications System": [
        "low voltage", "communications", "data", "telecom",
        "cabling", "fire alarm", "security", "cctv", "access control"
    ],

    "Site Utilities System": [
        "site utilities", "underground utilities",
        "water service", "sanitary service",
        "storm service", "gas service"
    ]
}


REGION_PROFILES = {
    "US": {
        "standard": "RSMeans",
        "typical_units": ["EA", "LF", "SF", "SY", "CY", "TON", "LS", "HR", "GAL"]
    },
    "COMMONWEALTH": {
        "standard": "NRM2/RICS",
        "typical_units": ["nr", "m", "m2", "m3", "t", "item", "sum", "hr", "l"]
    }
}


# -----------------------------
# DATA STRUCTURES
# -----------------------------

@dataclass
class OCRPage:
    page_num: int
    text: str

@dataclass
class Chunk:
    chunk_id: str
    page_start: int
    page_end: int
    token_est: int
    text: str

@dataclass
class SystemChunk:
    system_name: str
    referenced_text: str

# ============================================================
# Cache Builders
# ============================================================

def build_csi_candidates(price_data: List[Dict[str, Any]]) -> Dict[str, Dict[str, List]]:
    csi_map: Dict[str, Dict[str, List]] = {}

    for p in price_data:
        csi = p.get("CSI", "").strip()
        text = p.get("item", "").strip()
        if not csi or not text:
            continue

        if csi not in csi_map:
            csi_map[csi] = {
                "material": [],
                "labor": [],
                "equipment": []
            }

        if p.get("material_unit_cost", 0) > 0:
            csi_map[csi]["material"].append(p)
        if p.get("labor_rate", 0) > 0:
            csi_map[csi]["labor"].append(p)
        if p.get("equipment_rate", 0) > 0:
            csi_map[csi]["equipment"].append(p)

    return csi_map


def build_faiss_indices(csi_to_candidates: Dict[str, Dict[str, List]]) -> Dict:
    csi_to_index: Dict[str, Dict[str, Any]] = {}

    for csi, groups in csi_to_candidates.items():
        csi_to_index[csi] = {}

        for res, items in groups.items():
            if not items:
                continue

            texts = [p["item"] for p in items]
            embeddings = MATCH_MODEL.encode(texts).astype("float32")
            faiss.normalize_L2(embeddings)

            index = faiss.IndexFlatIP(embeddings.shape[1])
            index.add(embeddings)

            csi_to_index[csi][res] = {
                "index": index,
                "items": items
            }

    return csi_to_index


def load_or_build_cache() -> Tuple[Dict, Dict]:
    os.makedirs(CACHE_DIR, exist_ok=True)

    if os.path.exists(CANDIDATES_JSON) and os.path.exists(FAISS_PKL):
        try:
            print("🔍 Loading cache...")
            with open(CANDIDATES_JSON, "r", encoding="utf-8") as f:
                csi_to_candidates = json.load(f)
            with open(FAISS_PKL, "rb") as f:
                csi_to_index = pickle.load(f)
            return csi_to_candidates, csi_to_index
        except Exception:
            print("⚠️ Cache corrupted. Rebuilding...")

    print("⚙️ Building cache...")
    with open(PRICE_JSON, "r", encoding="utf-8") as f:
        price_data = json.load(f)

    csi_to_candidates = build_csi_candidates(price_data)
    csi_to_index = build_faiss_indices(csi_to_candidates)

    with open(CANDIDATES_JSON, "w", encoding="utf-8") as f:
        json.dump(csi_to_candidates, f)

    with open(FAISS_PKL, "wb") as f:
        pickle.dump(csi_to_index, f, protocol=pickle.HIGHEST_PROTOCOL)

    return csi_to_candidates, csi_to_index

# ============================================================
# Matching Logic
# ============================================================

def faiss_match(
    query: str,
    csi: str,
    resource: str,
    csi_to_index: Dict
) -> Tuple[Any, float, List[Any]]:

    bucket = csi_to_index.get(csi, {}).get(resource)
    if not bucket:
        return None, 0.0, []

    q_emb = MATCH_MODEL.encode([query]).astype("float32")
    faiss.normalize_L2(q_emb)

    scores, idxs = bucket["index"].search(q_emb, TOP_K)
    scores = scores[0]
    idxs = idxs[0]

    candidates = []
    final_scores = []

    # print("\n" + "=" * 80)
    # print(f"🔎 FAISS MATCH")
    # print(f"CSI      : {csi}")
    # print(f"Resource : {resource}")
    # print(f"Query    : {query}")
    # print("-" * 80)

    for rank, idx in enumerate(idxs):
        item = bucket["items"][idx]
        kw = keyword_similarity(query, item["item"])
        score = 0.8 * scores[rank] + 0.2 * kw
        candidates.append(item)
        final_scores.append(score)

        # print(
        #     f"[{rank}] "
        #     f"SEM={scores[rank]:.3f} "
        #     f"KW={kw:.3f} "
        #     f"SCORE={score:.3f} | "
        #     f"{item['item']}"
        # )

    best_idx = int(np.argmax(final_scores))
    best_score = float(final_scores[best_idx])
    best_item = candidates[best_idx]

    # print("-" * 80)
    # print(
    #     f"✅ BEST (FAISS) → "
    #     f"SCORE={best_score:.3f} | "
    #     f"{best_item['item']}"
    # )
    # print("=" * 80)
    return best_item, best_score, candidates


def llm_fallback(
    query: str,
    csi: str,
    candidates: List[Dict[str, Any]]
) -> Tuple[Any, float]:

    # print("⚠️ FAISS CONFIDENCE LOW → LLM FALLBACK")
    # print(f"Query: {query}")
    # print("Candidates sent to LLM:")
    # for i, c in enumerate(candidates):
    #     print(f"  [{i}] {c['item']}")

    options = [{"id": i, "item": c["item"]} for i, c in enumerate(candidates)]

    prompt = f"""
You are a construction estimator.

CSI: {csi}
Job Activity: "{query}"

Choose the BEST matching item.
Return JSON only.

Options:
{json.dumps(options, indent=2)}

{{"id": number, "confidence": number}}
"""

    resp = client.chat.completions.create(
        model="gpt-4.1-mini",
        temperature=0,
        messages=[{"role": "user", "content": prompt}]
    )

    try:
        data = json.loads(resp.choices[0].message.content)
        idx = int(data.get("id", -1))
        conf = float(data.get("confidence", 0))
        if idx < 0 or idx >= len(candidates):
            return None, conf

        # print(
        #     f"🤖 LLM SELECTED → "
        #     f"[{idx}] {candidates[idx]['item']} "
        #     f"(confidence={conf:.2f})"
        # )

        return candidates[idx], conf
    except Exception:
        # print("❌ LLM FAILED TO SELECT")
        return None, 0.0
        
# ===============================
# UTILS
# ===============================

def keyword_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    a_tokens = set(a.lower().split())
    b_tokens = set(b.lower().split())
    return len(a_tokens & b_tokens) / max(len(a_tokens), 1)

def print_step(title: str):
    print("\n" + "=" * 80)
    print(title)
    print("=" * 80)

def normalize_whitespace(s: str) -> str:
    s = s.replace("\x0c", " ")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def estimate_tokens(text: str) -> int:
    # Rough token estimate: ~4 chars/token typical for English.
    return max(1, math.ceil(len(text) / 4))

def is_valid_filter_prompt(filter_sentence: Optional[str]) -> bool:
    if not filter_sentence:
        return False
    s = filter_sentence.strip().lower()
    # Heuristic: must contain at least 2 words and a system hint word.
    if len(s.split()) < 2:
        return False

    return True
    # hint_words = ["only", "just", "focus", "include", "exclude", "roof", "hvac", "plumb",
    #               "elect", "fire", "stairs", "concrete", "masonry", "finishes", "openings"]
    # return any(w in s for w in hint_words)

def fuzzy_merge_names(names: List[str], cutoff: float = 0.82) -> List[str]:
    merged = []
    for n in names:
        n_clean = n.strip()
        if not n_clean:
            continue
        placed = False
        for i, m in enumerate(merged):
            if difflib.SequenceMatcher(None, n_clean.lower(), m.lower()).ratio() >= cutoff:
                # Keep shorter / "brief" name preference
                merged[i] = min([m, n_clean], key=len)
                placed = True
                break
        if not placed:
            merged.append(n_clean)
    return merged

def normalize_system_name(name: str) -> str:
    lname = name.lower()
    for canonical, aliases in SYSTEM_ONTOLOGY.items():
        if any(a in lname for a in aliases):
            return canonical
    return name.strip().title()

def brief_system_name(name: str) -> str:
    # Enforce brief naming (remove "system" repetition, trim).
    name = name.strip()
    name = re.sub(r"\bsystems?\b", "", name, flags=re.I).strip()
    name = re.sub(r"\s{2,}", " ", name)
    # Title-case but keep acronyms.
    return name[:60]

def clamp_allowed_category(category: str) -> str:
    if category in ALLOWED_CATEGORIES:
        return category
    # best-effort fuzzy match
    best = None
    best_score = 0.0
    for c in ALLOWED_CATEGORIES:
        score = difflib.SequenceMatcher(None, category.lower(), c.lower()).ratio()
        if score > best_score:
            best_score = score
            best = c
    return best if best_score >= 0.7 else "General Requirements"


# -----------------------------
# OCR: PDF -> text per page
# -----------------------------

def ocr_pdf(
    pdf_path: str,
    dpi: int = 200,
    poppler_path: Optional[str] = None,
    tesseract_cmd: Optional[str] = None,
    lang: str = "eng"
) -> List[OCRPage]:
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

    images = convert_from_path(pdf_path, dpi=dpi, poppler_path=poppler_path)
    pages: List[OCRPage] = []

    for i, img in enumerate(images, start=1):
        text = pytesseract.image_to_string(img, lang=lang)
        text = normalize_whitespace(text)
        pages.append(OCRPage(page_num=i, text=text))
        print(f"OCR page {i}/{len(images)}: {len(text)} chars")

    return pages


# -----------------------------
# HYBRID CHUNKING (page + token)
# Optimized for assemblies: keeps "notes" blocks together when possible.
# -----------------------------

NOTES_HEADINGS = [
    "GENERAL NOTES", "SPECIAL NOTES", "GENERAL", "NOTES",
    "SPECIFICATIONS", "SCOPE", "LEGEND", "ABBREVIATIONS"
]

def detect_notes_bias(text: str) -> bool:
    up = text.upper()
    return any(h in up for h in NOTES_HEADINGS)

def hybrid_chunk_pages(
    ocr_pages: List[OCRPage],
    max_tokens: int = 1400,
    hard_max_tokens: int = 1800,
    min_tokens: int = 350
) -> List[Chunk]:
    chunks: List[Chunk] = []

    buf_pages: List[OCRPage] = []
    buf_texts: List[str] = []
    buf_tokens = 0

    def flush():
        nonlocal buf_pages, buf_texts, buf_tokens
        if not buf_pages:
            return
        text = "\n\n".join(buf_texts).strip()
        if not text:
            buf_pages, buf_texts, buf_tokens = [], [], 0
            return
        chunk = Chunk(
            chunk_id=f"p{buf_pages[0].page_num}-p{buf_pages[-1].page_num}",
            page_start=buf_pages[0].page_num,
            page_end=buf_pages[-1].page_num,
            token_est=estimate_tokens(text),
            text=text
        )
        chunks.append(chunk)
        buf_pages, buf_texts, buf_tokens = [], [], 0

    for p in ocr_pages:
        p_tokens = estimate_tokens(p.text)
        # If a single page is too large, split within page by paragraphs.
        if p_tokens > hard_max_tokens:
            flush()
            paras = [x.strip() for x in re.split(r"\n\s*\n", p.text) if x.strip()]
            cur = []
            cur_t = 0
            for para in paras:
                t = estimate_tokens(para)
                if cur and cur_t + t > max_tokens:
                    text = "\n\n".join(cur)
                    chunks.append(Chunk(
                        chunk_id=f"p{p.page_num}-split{len(chunks)+1}",
                        page_start=p.page_num, page_end=p.page_num,
                        token_est=estimate_tokens(text),
                        text=text
                    ))
                    cur, cur_t = [], 0
                cur.append(para)
                cur_t += t
            if cur:
                text = "\n\n".join(cur)
                chunks.append(Chunk(
                    chunk_id=f"p{p.page_num}-split{len(chunks)+1}",
                    page_start=p.page_num, page_end=p.page_num,
                    token_est=estimate_tokens(text),
                    text=text
                ))
            continue

        # Notes bias: prefer keeping notes with adjacent page if still within max_tokens.
        bias = detect_notes_bias(p.text)

        # Normal aggregation
        if buf_pages and (buf_tokens + p_tokens > max_tokens) and (buf_tokens >= min_tokens):
            flush()

        # If notes-bias and small overflow, allow it unless hard limit
        if buf_pages and bias and (buf_tokens + p_tokens <= hard_max_tokens):
            buf_pages.append(p)
            buf_texts.append(f"[PAGE {p.page_num}]\n{p.text}")
            buf_tokens += p_tokens
            continue

        # If still too big, flush first then start new
        if buf_pages and (buf_tokens + p_tokens > hard_max_tokens):
            flush()

        buf_pages.append(p)
        buf_texts.append(f"[PAGE {p.page_num}]\n{p.text}")
        buf_tokens += p_tokens

    flush()
    return chunks


# -----------------------------
# GPT HELPERS
# -----------------------------

def gpt_json(system_prompt: str, user_prompt: str) -> Any:
    resp = client.responses.create(
        model=MODEL,
        input=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        # Responses API: ask for JSON
        text={"format": {"type": "json_object"}},
    )
    return json.loads(resp.output_text)

def build_region_instruction(region: str) -> str:
    region = region.upper().strip()
    profile = REGION_PROFILES.get(region, REGION_PROFILES["US"])
    return (
        f"Region standard: {profile['standard']}. "
        f"Prefer these units: {', '.join(profile['typical_units'])}. "
        "Use region-appropriate unit conventions. DON't include another measure units."
    )


# -----------------------------
# SYSTEM CLASSIFICATION + MERGE
# -----------------------------

SYSTEM_CLASSIFIER_SYS = f"""You are a construction estimator assistant.
Task: From OCR text, identify building systems mentioned (assemblies).

SYSTEM RESTRICTION (MANDATORY)
You may ONLY output systems that exactly match one of the following
canonical system names:
{list(SYSTEM_ONTOLOGY.keys())}

CRITICAL OWNERSHIP RULE:
- Each OCR sentence or paragraph may belong to ONE system only.
- If scope overlaps, assign the text to the system that CONSUMES the work.
- Do NOT duplicate the same OCR text across multiple systems.

Return JSON only.
"""

def classify_systems_in_chunk(chunk: Chunk) -> List[Dict[str, str]]:
    user = f"""
Extract top-level systems from the following OCR chunk.

Rules:
- Return a list of systems with brief names.
- Prefer common systems such as Roofing, Stairs, HVAC, Plumbing, Electrical, Fire Protection, etc.
- An item belongs to the system that consumes it.
- Ignore any notes, instructions, disclaimers or code references
- Include "General Requirements" only if scope/admin requirements are explicit.
- M.H is manhole and manhole MUST be in Manhole and Structure system, not in Sanitary Sewer System. Item will be Manhole Installation (Including Frame and Cover).

For each system include:
  - system_name (brief)
  - referenced_text (direct quotes/snippets from the chunk that justify it; keep concise):
    - Include ALL OCR sentences or paragraphs from the chunk that are directly relevant to this system's scope, components, quantities, materials, dimensions, or installation requirements.
    - Preserve original wording exactly (verbatim OCR text).
    - Combine related sentences into coherent blocks when they belong together.
    - EXCLUDE unrelated notes, boilerplate, legends, code citations, or instructions for other systems.
    - Do NOT summarize or paraphrase.
    - Do NOT include the same OCR text under multiple systems unless it clearly applies to more than one system.

OCR CHUNK {chunk.chunk_id}:
{chunk.text}
"""
    data = gpt_json(
        system_prompt=SYSTEM_CLASSIFIER_SYS,
        user_prompt=user
    )
    # expected: {"systems":[{"system_name":"HVAC","referenced_text":"..."}]}
    systems = data.get("systems", [])
    out = []
    for s in systems:
        name = normalize_system_name(str(s.get("system_name", "")).strip())
        ref = str(s.get("referenced_text", "")).strip()
        if name and ref:
            out.append({"system_name": name, "referenced_text": ref})
    return out

def merge_systems_across_chunks(per_chunk_systems: Dict[str, List[Dict[str, str]]]) -> Dict[str, List[Tuple[str, str]]]:
    """
    Returns:
      {system_name: [(chunk_id, referenced_text), ...]}
    """
    all_names = []
    for chunk_id, systems in per_chunk_systems.items():
        for s in systems:
            all_names.append(s["system_name"])

    merged_names = fuzzy_merge_names(all_names, cutoff=0.82)
    # Map each original name to merged representative
    rep_map = {}
    for n in set(all_names):
        best = None
        best_score = -1.0
        for rep in merged_names:
            score = difflib.SequenceMatcher(None, n.lower(), rep.lower()).ratio()
            if score > best_score:
                best = rep
                best_score = score
        rep_map[n] = best if best else n

    merged: Dict[str, List[Tuple[str, str]]] = {}
    for chunk_id, systems in per_chunk_systems.items():
        for s in systems:
            rep = rep_map.get(s["system_name"], s["system_name"])
            merged.setdefault(rep, []).append((chunk_id, s["referenced_text"]))

    # De-dup referenced_text lines
    for sysname in list(merged.keys()):
        seen = set()
        uniq = []
        for cid, ref in merged[sysname]:
            key = (cid, ref)
            if key in seen:
                continue
            seen.add(key)
            uniq.append((cid, ref))
        merged[sysname] = uniq

    return merged

def build_system_chunks(merged_system_refs: Dict[str, List[Tuple[str, str]]]) -> List[SystemChunk]:
    system_chunks: List[SystemChunk] = []
    for sysname, refs in merged_system_refs.items():
        collected = []
        for cid, ref in refs:
            collected.append(f"[{cid}] {ref}")
        # system_chunks.append(SystemChunk(system_name=normalize_system_name(sysname),
        #                                  referenced_text="\n".join(collected).strip()))
        system_chunks.append(SystemChunk(system_name=sysname,
                                         referenced_text="\n".join(collected).strip()))
    # stable sort
    system_chunks.sort(key=lambda x: x.system_name.lower())
    return system_chunks


# -----------------------------
# FILTERING SYSTEM CHUNKS BY USER FILTER SENTENCE
# (Only used if valid filter prompt)
# -----------------------------

FILTER_SYS = """You are a classifier that selects which system names to process.
Given a user filter sentence and a list of system names, choose which systems match.
Return JSON only: {"selected_systems":[...]}.
"""

def filter_systems_with_gpt(
    filter_sentence: str,
    system_names: List[str]
) -> List[str]:
    user = f"""
User filter sentence:
{filter_sentence}

System names:
{json.dumps(system_names, indent=2)}

Select only the systems that should be processed based on the filter sentence.
"""
    data = gpt_json(FILTER_SYS, user)
    selected = data.get("selected_systems", [])
    selected = [s for s in selected if s in system_names]
    return selected

# -----------------------------
# LINE ITEM EXTRACTION (CSI + category + qty + unit)
# -----------------------------

ITEM_EXTRACTOR_SYS = """You are a construction cost estimator.
Extract estimatable line items from system text, aligned to MasterFormat/CSI and estimate labor productivity in labor-hours per unit and equipment productivity in hours per unit.
Output JSON only. Use the allowed categories exactly.
Quantity of elements like manhole (M.H, M.H.#1), cleanout, valve, etc should be counted as individual units.
If quantity is missing, choose a reasonable default dimension/quantity assumption.
Do not invent scope not supported by the text; but you may infer standard components when notes imply them.
Ensure CSI division and section are plausible.

INVALID RULES:
- M.H is not "main panel". M.H is manhole.

Schema:
{
  "items":[
    {"DIV":"##","CSI":"## ## ##","Category":"...allowed...","Item":"...","quantity":number,"unit":"...","labor_hours_per_unit":number,"equipment_hours_per_unit":number}
  ]
}
"""

def extract_items_for_system(
    system_name: str,
    system_text: str,
    allowed_categories: List[str],
    region: str,
    user_prompt_extra: str
) -> List[Dict[str, Any]]:

    region_instruction = build_region_instruction(region)

    user = f"""
System: {system_name}

Allowed categories (must choose one exactly):
{json.dumps(allowed_categories, indent=2)}

Region:
{region_instruction}

Estimator instruction (Just filter):
{user_prompt_extra}

Now extract line items from the referenced text below.

Referenced Text:
{system_text}
"""
    data = gpt_json(
        system_prompt=ITEM_EXTRACTOR_SYS,
        user_prompt=user
    )
    items = data.get("items", [])
    cleaned = []
    for it in items:
        div = str(it.get("DIV", "")).strip()
        csi = str(it.get("CSI", "")).strip()
        cat = clamp_allowed_category(str(it.get("Category", "")).strip())
        item = str(it.get("Item", "")).strip()
        qty = it.get("quantity", None)
        unit = str(it.get("unit", "")).strip()
        labor_hours_per_unit = it.get("labor_hours_per_unit", 0)
        equipment_hours_per_unit = it.get("equipment_hours_per_unit", 0)

        # basic cleanup
        if not item or not unit:
            continue
        try:
            qty = float(qty)
        except Exception:
            # allow missing -> skip; model should provide default, but just in case:
            continue

        cleaned.append({
            "DIV": div,
            "CSI": csi,
            "Category": cat,
            "Item": item,
            "quantity": qty,
            "unit": unit,
            "labor_hours_per_unit": labor_hours_per_unit,
            "equipment_hours_per_unit": equipment_hours_per_unit
        })
    return cleaned

# -----------------------------
# Cost Estimate (material costs + labor + equipment)
# -----------------------------

def estimate_costs_for_items(
    system_to_items: Dict[str, List[Dict[str, Any]]],
    region: str,
    csi_to_index: Dict
) -> Dict[str, List[Dict[str, Any]]]:

    cost_items = {}

    for system_name, items in system_to_items.items():
        enriched = []

        for item in items:
            qty = item.get("quantity", 0)

            item["L.Hrs"] = item.get("labor_hours_per_unit", 0) * qty
            item["E.Hrs"] = item.get("equipment_hours_per_unit", 0) * qty

            for res, rate_key, rate_out, total_out in [
                ("material", "material_unit_cost", "M.Cost", "T.Mat"),
                ("labor", "labor_rate", "L.Rate", "T.Labor"),
                ("equipment", "equipment_rate", "E.Rate", "T.Equip"),
            ]:
                best, score, topk = faiss_match(
                    item["Item"], item["CSI"], res, csi_to_index
                )

                source = "faiss"
                if score < FAISS_THRESHOLD:
                    llm_best, llm_score = llm_fallback(
                        item["Item"], item["CSI"], topk
                    )
                    if llm_best:
                        best = llm_best
                        score = llm_score
                        source = "llm"

                rate = best.get(rate_key, 0) if best else 0
                item[rate_out] = rate
                item[total_out] = rate * qty

                item.setdefault(res, {})
                item[res]["matched_price"] = best
                item[res]["similarity"] = round(score, 3)
                item[res]["source"] = source

            enriched.append(item)

        cost_items[system_name] = enriched

    return cost_items

# -----------------------------
# EXCEL EXPORT (one sheet, system name as merged row)
# -----------------------------

def export_to_excel(
    out_path: str,
    system_to_items: Dict[str, List[Dict[str, Any]]]
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    headers = ["DIV", "CSI", "Category", "Item", "quantity", "unit", "M.Cost", "T.Mat", "L.Rate", "L.Hrs", "T.Labor", "E.Rate", "E.Hrs", "T.Equip", "Sub Total"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    system_fill = PatternFill("solid", fgColor="D9E1F2")
    system_font = Font(bold=True)

    for system_name in sorted(system_to_items.keys(), key=lambda x: x.lower()):
        items = system_to_items[system_name]
        if not items:
            continue

        # system merged row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        c = ws.cell(row=row, column=1, value=system_name)
        c.fill = system_fill
        c.font = system_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        # items
        for it in items:
            row_total = sum([
                it.get("T.Mat", 0),
                it.get("T.Labor", 0),
                it.get("T.Equip", 0)
            ])
            it["Sub Total"] = row_total
            ws.append([it.get(h, "") for h in headers])
            row += 1

    # column widths
    widths = {
        "A": 8,   # DIV
        "B": 12,  # CSI
        "C": 30,  # Category
        "D": 60,  # Item
        "E": 12,  # quantity
        "F": 10,  # unit
        "G": 12,  # M.Cost
        "H": 12,  # T.Mat
        "I": 12,  # L.Rate
        "J": 12,  # L.Hrs
        "K": 12,  # T.Labor
        "L": 12,  # E.Rate
        "M": 12,  # E.Hrs
        "N": 12,  # T.Equip
        "O": 12   # Sub Total
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    wb.save(out_path)


# -----------------------------
# MAIN PIPELINE
# -----------------------------

def run(
    pdf_path: str,
    out_xlsx: str,
    region: str = "US",
    user_filter_sentence: Optional[str] = None,
    poppler_path: Optional[str] = None,
    tesseract_cmd: Optional[str] = None,
    ocr_lang: str = "eng"
):

    # --------------------------------------------------
    # Preprocess: group by CSI and pre-embed items
    # --------------------------------------------------

    csi_to_candidates, csi_to_index = load_or_build_cache()


    # --------------------------------------------------
    # Preprocess: OCR + Chunking + System Classification + Merging
    # --------------------------------------------------
    print_step("1) OCR PDF")
    ocr_pages = ocr_pdf(
        pdf_path=pdf_path,
        dpi=200,
        poppler_path=poppler_path,
        tesseract_cmd=tesseract_cmd,
        lang=ocr_lang
    )
    print(f"Total pages OCR'd: {len(ocr_pages)}")

    print_step("2) Hybrid chunking (page + token)")
    chunks = hybrid_chunk_pages(ocr_pages, max_tokens=1400, hard_max_tokens=1800, min_tokens=350)
    print(f"Total chunks: {len(chunks)}")
    for ch in chunks[:5]:
        print(f"Chunk {ch.chunk_id}: pages {ch.page_start}-{ch.page_end}, token_est={ch.token_est}")
    if len(chunks) > 5:
        print("...")

    print_step("3) Classify systems per chunk (GPT)")
    per_chunk_systems: Dict[str, List[Dict[str, str]]] = {}
    for i, ch in enumerate(chunks, start=1):
        print(f"Classifying chunk {i}/{len(chunks)}: {ch.chunk_id}")
        systems = classify_systems_in_chunk(ch)
        per_chunk_systems[ch.chunk_id] = systems
        print(f"  Systems found: {len(systems)}")
        for s in systems[:6]:
            print(f"   - {s['system_name']}: {s['referenced_text'][:90]}")

    print_step("4) Merge similar systems across chunks")
    merged_system_refs = merge_systems_across_chunks(per_chunk_systems)
    system_chunks = build_system_chunks(merged_system_refs)
    print(f"Merged systems: {len(system_chunks)}")
    for sc in system_chunks:
        print(f" - {sc.system_name} (refs: {len(merged_system_refs.get(sc.system_name, []))})")

    print_step("5) Optional user filter (only if valid)")
    system_names = [sc.system_name for sc in system_chunks]
    if is_valid_filter_prompt(user_filter_sentence):
        print(f"Filter sentence is valid. Applying: {user_filter_sentence}")
        selected = filter_systems_with_gpt(user_filter_sentence, system_names)
        selected_set = set(selected)
        system_chunks = [sc for sc in system_chunks if sc.system_name in selected_set]
        print(f"Systems selected: {len(system_chunks)}")
        for s in selected:
            print(f" - {s}")
    else:
        print("No valid filter sentence provided; processing all systems.")

    print_step("6) Extract line items per system (GPT) + MasterFormat/CSI + allowed categories")
    system_to_items: Dict[str, List[Dict[str, Any]]] = {}
    for i, sc in enumerate(system_chunks, start=1):
        print(f"Extracting items {i}/{len(system_chunks)}: {sc.system_name}")
        items = extract_items_for_system(
            system_name=sc.system_name,
            system_text=sc.referenced_text,
            allowed_categories=ALLOWED_CATEGORIES,
            region=region,
            user_prompt_extra=user_filter_sentence
        )
        system_to_items[sc.system_name] = items
        print(f"  Items extracted: {len(items)}")
        for it in items[:5]:
            print(f"   - {it['CSI']} | {it['Category']} | {it['Item']} | {it['quantity']} {it['unit']}")

    print_step("7) Estimate costs for items (GPT)")
    cost_items = estimate_costs_for_items(system_to_items, region, csi_to_index)

    print_step("8) Export to Excel (one sheet, system name as merged row)")
    export_to_excel(out_xlsx, cost_items)
    print(f"Excel saved: {out_xlsx}")


if __name__ == "__main__":
    # Example usage:
    #   python estimate_from_pdf.py
    #
    # Configure via environment variables or edit below.
    PDF_PATH = os.getenv("PDF_PATH", "Mikie Building Plans_ Jamaica.pdf")
    OUT_XLSX = os.getenv("OUT_XLSX", "estimate.xlsx")

    REGION = os.getenv("REGION", "US")  # "US" or "COMMONWEALTH"
    USER_FILTER = None #"Please estimate only manholes." # os.getenv("USER_FILTER", "").strip() or None

    POPPLER_PATH = os.getenv("POPPLER_PATH")  # optional (Windows)
    TESSERACT_CMD = os.getenv("TESSERACT_CMD")  # optional (Windows)
    OCR_LANG = os.getenv("OCR_LANG", "eng")

    MODEL = os.getenv("MODEL", "gpt-5.2")

    run(
        pdf_path=PDF_PATH,
        out_xlsx=OUT_XLSX,
        region=REGION,
        user_filter_sentence=USER_FILTER,
        poppler_path=POPPLER_PATH,
        tesseract_cmd=TESSERACT_CMD,
        ocr_lang=OCR_LANG
    )